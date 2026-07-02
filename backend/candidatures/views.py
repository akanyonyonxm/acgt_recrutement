"""API DRF — pilotage des statuts de dossier.

Le changement de statut ne se fait jamais par écriture directe du champ : il
passe par des actions dédiées qui valident la transition (via le modèle) et
contrôlent le rôle de l'utilisateur.
"""

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import transaction
from django.db.models import Count, Exists, OuterRef, Q
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.throttling import AnonRateThrottle

from . import roles
from .models import (
    EXTENSIONS_AUTORISEES,
    TAILLE_MAX_PIECE,
    AffectationEvaluateur,
    AppelCandidature,
    ControleCritere,
    CritereValidation,
    DocumentReclamation,
    Dossier,
    EmailQueue,
    Evaluation,
    HistoriqueStatut,
    ListeEligibilite,
    PieceJointe,
    Poste,
    ReclamationEligibilite,
    Recours,
    RetenuDefinitif,
    TypePiece,
)
from .pagination import PaginationPublique, PaginationStandard
from .permissions import EstAdminOuLectureSeule
from .serializers import (
    AffectationSerializer,
    AppelCandidatureSerializer,
    ChangementStatutSerializer,
    CritereValidationSerializer,
    DossierListeSerializer,
    DossierSerializer,
    EligibiliteAdminSerializer,
    EligibilitePubliqueSerializer,
    EvaluationSerializer,
    HistoriqueStatutSerializer,
    ModificationIdentiteSerializer,
    ModificationNomEligibiliteSerializer,
    PieceJointeSerializer,
    PieceJointeUploadSerializer,
    PosteSerializer,
    ReclamationAdminSerializer,
    ReclamationCreationSerializer,
    RecoursAdminSerializer,
    RecoursCreationSerializer,
    RecoursModificationSerializer,
    RetenuDefinitifSerializer,
    RetenuSupplementSerializer,
    RetenuPubliqueSerializer,
    TypePieceSerializer,
)
from .services.email import EmailError, envoyer_email
from .services.import_eligibilite import ImportEligibiliteErreur, importer_eligibles
from .utils import normaliser_texte, tokens_recherche

User = get_user_model()


class TypePieceViewSet(viewsets.ReadOnlyModelViewSet):
    """Référentiel des types de pièce (lecture seule ; édition en Django Admin)."""

    queryset = TypePiece.objects.filter(actif=True)
    serializer_class = TypePieceSerializer
    permission_classes = [AllowAny]


class PosteViewSet(viewsets.ReadOnlyModelViewSet):
    """Référentiel des postes/fonctions (lecture seule ; édition en Django Admin)."""

    queryset = Poste.objects.filter(actif=True)
    serializer_class = PosteSerializer
    permission_classes = [AllowAny]


class EligibiliteViewSet(viewsets.ReadOnlyModelViewSet):
    """Liste d'éligibilité, en lecture seule, avec recherche tolérante (?q=).

    - Public : uniquement les personnes publiées, exposées en NOM/POSTNOM/PRÉNOM.
    - Admin  : toute la liste, tous les champs (référence interne incluse), pour
      la validation des dossiers.

    L'édition se fait dans Django Admin (et via l'import Excel).
    """

    permission_classes = [AllowAny]
    pagination_class = PaginationPublique

    def _staff(self):
        return roles.acces_backoffice(self.request.user)

    def get_serializer_class(self):
        if self._staff():
            return EligibiliteAdminSerializer
        return EligibilitePubliqueSerializer

    def get_queryset(self):
        qs = ListeEligibilite.objects.all()
        if not self._staff():
            qs = qs.filter(est_publie=True)
        # Recherche tolérante : chaque mot de la requête doit être contenu dans
        # le texte normalisé (ordre indifférent, insensible aux accents/casse).
        for token in tokens_recherche(self.request.query_params.get('q', '')):
            qs = qs.filter(texte_recherche__contains=token)
        return qs

    @action(detail=True, methods=['patch'], url_path='nom')
    def modifier_nom(self, request, pk=None):
        """Corrige le NOM (nom/postnom/prénom) d'une personne de la liste.

        Réservé aux administrateurs et correcteurs. Le **code n'est pas
        modifiable** ici (identifiant stable). `texte_recherche` est recalculé
        (les correspondances avec les dossiers restent cohérentes).
        """
        if not (roles.est_admin(request.user) or roles.est_correcteur(request.user)):
            raise PermissionDenied(
                "Seuls les administrateurs et correcteurs peuvent corriger un nom."
            )
        ligne = self.get_object()
        serializer = ModificationNomEligibiliteSerializer(ligne, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(EligibiliteAdminSerializer(ligne).data)

    @action(detail=False, methods=['get'], url_path='verifier-code')
    def verifier_code(self, request):
        """Vérifie qu'un code figure sur la liste publiée (aide à la saisie).

        Renvoie le nom tel qu'écrit sur la liste, à titre de repère visuel
        uniquement : le code seul fait foi, aucune comparaison de noms n'est
        faite et rien n'est bloquant. N'expose que des données déjà publiques
        (la liste des éligibles affiche code + noms).
        """
        code = (request.query_params.get('code') or '').strip()
        if not code:
            return Response({'trouve': False, 'multiple': False, 'ligne': None})
        lignes = list(
            ListeEligibilite.objects.filter(est_publie=True, code__iexact=code)[:2]
        )
        if not lignes:
            return Response({'trouve': False, 'multiple': False, 'ligne': None})
        if len(lignes) > 1:
            # Code ambigu (présent sur plusieurs lignes) : reconnu, mais sans
            # nom affichable ni rattachement automatique possible.
            return Response({'trouve': True, 'multiple': True, 'ligne': None})
        ligne = lignes[0]
        return Response({
            'trouve': True, 'multiple': False,
            'ligne': {
                'id': ligne.id, 'code': ligne.code, 'nom': ligne.nom,
                'postnom': ligne.postnom, 'prenom': ligne.prenom,
            },
        })

    @action(detail=False, methods=['post'], url_path='importer',
            parser_classes=[MultiPartParser, FormParser])
    def importer(self, request):
        """Importe un classeur Excel d'éligibles (admin).

        Champs multipart : `fichier` (.xlsx), `remplacer` (bool : vider d'abord),
        `publier` (bool : publier les lignes importées). Renvoie le récapitulatif.
        """
        if not roles.est_admin(request.user):
            raise PermissionDenied("Réservé aux administrateurs.")

        fichier = request.FILES.get('fichier')
        if not fichier:
            raise ValidationError({'fichier': "Aucun fichier fourni."})
        if not fichier.name.lower().endswith('.xlsx'):
            raise ValidationError({'fichier': "Format attendu : .xlsx"})

        def vrai(v):
            return str(v).lower() in ('1', 'true', 'on', 'oui')

        try:
            resultat = importer_eligibles(
                fichier,
                remplacer=vrai(request.data.get('remplacer')),
                publier=vrai(request.data.get('publier')),
            )
        except ImportEligibiliteErreur as exc:
            raise ValidationError({'detail': str(exc)})
        return Response(resultat)

    @action(detail=False, methods=['get'], url_path='modele')
    def modele(self, request):
        """Télécharge un modèle Excel vierge (en-têtes attendus) pour l'import."""
        if not roles.est_admin(request.user):
            raise PermissionDenied("Réservé aux administrateurs.")

        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Éligibles'
        entetes = ['code', 'nom', 'postnom', 'prenom', 'type', 'annee', 'reference']
        ws.append(entetes)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # Ligne-guide SANS nom : ignorée à l'import (« nom » obligatoire), mais
        # montre le format attendu sans risquer d'importer une fausse personne.
        # type et reference laissés vides (facultatifs).
        ws.append(['ACGT-001', '', 'Mukendi', 'Jean', '', 2021, ''])
        for i, largeur in enumerate([14, 18, 18, 18, 16, 8, 16], start=1):
            ws.column_dimensions[get_column_letter(i)].width = largeur

        reponse = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        reponse['Content-Disposition'] = 'attachment; filename="modele_eligibles.xlsx"'
        wb.save(reponse)
        return reponse


def _lettre_salle(n):
    """0 -> A, 1 -> B, … 25 -> Z, 26 -> AA (libellés de salle, style Excel)."""
    s = ''
    n += 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _docs_retenu_definitif(e):
    """Documents (pièces dossier + justificatifs réclamation) rattachés à une
    entrée définitive, pour vérifier l'identité (ex. date de naissance)."""
    dossiers, reclams = [], []
    if e.dossier_id and e.dossier:
        dossiers.append(e.dossier)
        reclams += list(e.dossier.reclamation_origine.all())
    if e.recours_id and e.recours:
        if e.recours.dossier_id and e.recours.dossier:
            dossiers.append(e.recours.dossier)
            reclams += list(e.recours.dossier.reclamation_origine.all())
        if e.recours.reclamation_id and e.recours.reclamation:
            reclams.append(e.recours.reclamation)
    docs = []
    for d in dossiers:
        for p in d.pieces.all():
            docs.append({'libelle': p.type_piece.libelle, 'nom_original': p.nom_original,
                         'url': f'/api/dossiers/{d.id}/pieces/{p.id}/telecharger/'})
    for rc in reclams:
        for doc in rc.documents.all():
            docs.append({'libelle': doc.get_type_display(), 'nom_original': doc.nom_original,
                         'url': f'/api/reclamations/{rc.id}/documents/{doc.id}/'})
    return docs


def _domaine_recours(r):
    """Libellé du domaine (poste) effectif d'un recours : la correction admin
    (`r.poste`) si présente, sinon le domaine hérité de la source (dossier ou
    réclamation liée). Source unique de vérité, utilisée pour la liste
    définitive comme pour l'affichage back-office."""
    if r.poste_id:
        return r.poste.libelle
    if r.dossier_id and r.dossier.poste_id:
        return r.dossier.poste.libelle
    if r.reclamation_id and r.reclamation.poste_id:
        return r.reclamation.poste.libelle
    return ''


def _sources_liste_definitive(appel):
    """Liste combinée de la liste DÉFINITIVE : retenus publiés (dossiers RETENU)
    + recours VALIDÉS rattachés à cet appel, dédupliqués par identité normalisée
    (un retenu prime sur son éventuel recours). Triée par nom. Renvoie des dicts."""
    vus = set()
    out = []
    for d in (appel.dossiers.filter(statut=Dossier.Statut.RETENU)
              .select_related('poste')):
        tr = d.texte_recherche or normaliser_texte(f'{d.nom} {d.postnom} {d.prenom}')
        if tr in vus:
            continue
        vus.add(tr)
        out.append({
            'texte_recherche': tr, 'nom': d.nom, 'postnom': d.postnom, 'prenom': d.prenom,
            'poste_libelle': d.poste.libelle if d.poste else '',
            'origine': RetenuDefinitif.Origine.LISTE, 'dossier_id': d.id, 'recours_id': None,
        })
    rec_qs = (Recours.objects.filter(statut=Recours.Statut.VALIDE)
              .filter(Q(dossier__appel=appel) | Q(reclamation__appel=appel))
              .select_related('poste', 'dossier__poste', 'reclamation__poste'))
    for r in rec_qs:
        tr = normaliser_texte(f'{r.nom} {r.postnom} {r.prenom}')
        if tr in vus:
            continue
        vus.add(tr)
        poste = _domaine_recours(r)
        out.append({
            'texte_recherche': tr, 'nom': r.nom, 'postnom': r.postnom, 'prenom': r.prenom,
            'poste_libelle': poste,
            'origine': RetenuDefinitif.Origine.RECOURS, 'dossier_id': None, 'recours_id': r.id,
        })
    # Ajouts supplémentaires (liste décidée hors plateforme, sans dossier/recours
    # source) : déjà figés dans RetenuDefinitif. On les réinjecte ici pour qu'ils
    # apparaissent dans l'aperçu back-office et le PDF (le public et l'Excel lisent
    # la table directement). _generer_liste_definitive les ignore (déjà présents).
    # Masqués si l'appel a désactivé l'affichage des suppléments.
    sup_qs = (appel.retenus_definitifs.filter(origine=RetenuDefinitif.Origine.SUPPLEMENT)
              if appel.afficher_supplements_definitif else appel.retenus_definitifs.none())
    for e in sup_qs:
        tr = e.texte_recherche or normaliser_texte(f'{e.nom} {e.postnom} {e.prenom}')
        if tr in vus:
            continue
        vus.add(tr)
        out.append({
            'texte_recherche': tr, 'nom': e.nom, 'postnom': e.postnom, 'prenom': e.prenom,
            'poste_libelle': e.poste_libelle,
            'origine': RetenuDefinitif.Origine.SUPPLEMENT, 'dossier_id': None, 'recours_id': None,
        })
    out.sort(key=lambda s: (s['nom'].lower(), s['postnom'].lower(), s['prenom'].lower()))
    return out


def _generer_liste_definitive(appel):
    """Génère/complète les entrées figées de la liste définitive. Les codes déjà
    attribués sont CONSERVÉS (stables) ; les nouvelles personnes prennent les
    codes suivants. Renvoie le nombre de nouvelles entrées créées."""
    existants = {e.texte_recherche: e for e in appel.retenus_definitifs.all()}
    codes = [int(e.code) for e in existants.values() if e.code.isdigit()]
    prochain = (max(codes) + 1) if codes else 1
    crees = 0
    for s in _sources_liste_definitive(appel):
        if s['texte_recherche'] in existants:
            continue
        RetenuDefinitif.objects.create(
            appel=appel, code=f'{prochain:04d}',
            nom=s['nom'], postnom=s['postnom'], prenom=s['prenom'],
            poste_libelle=s['poste_libelle'], origine=s['origine'],
            dossier_id=s['dossier_id'], recours_id=s['recours_id'],
        )
        prochain += 1
        crees += 1
    return crees


def _resultats_emails(appel):
    """Consolide les e-mails de résultat à envoyer pour un appel, UN par personne
    (dédup par nom normalisé) :

    - **admis** = personnes de la liste DÉFINITIVE (RetenuDefinitif) ;
    - **non retenu** = personnes décidées défavorablement et PAS sur la définitive,
      d'après leur **dernier traitement** (recours / réclamation / dossier, celui
      dont la date de décision est la plus récente) + le motif correspondant.

    Renvoie (admis: list, non_retenus: list) de dicts."""
    # 1) Admis = liste définitive (déjà dédupliquée, code stable)
    admis = {}
    for e in appel.retenus_definitifs.select_related('dossier', 'recours').all():
        cle = e.texte_recherche or normaliser_texte(f'{e.nom} {e.postnom} {e.prenom}')
        email = ''
        if e.dossier_id and e.dossier:
            email = e.dossier.email
        elif e.recours_id and e.recours:
            email = e.recours.email
        admis[cle] = {
            'type': 'admis', 'cle': cle, 'nom': e.nom, 'postnom': e.postnom, 'prenom': e.prenom,
            'email': (email or '').strip(), 'code': e.code, 'ville': e.get_ville_examen_display(),
        }
    noms_admis = set(admis)

    # 2) Non retenus = dernier traitement défavorable des personnes hors définitive
    cands = {}

    def maj(cle, date, type_, motif, nom, postnom, prenom, email):
        if not cle or cle in noms_admis or date is None:
            return
        cur = cands.get(cle)
        if cur is None or date > cur['date']:
            cands[cle] = {'date': date, 'dernier_traitement': type_, 'motif': (motif or '').strip(),
                          'nom': nom, 'postnom': postnom, 'prenom': prenom, 'email': (email or '').strip()}

    for d in (appel.dossiers
              .filter(statut__in=[Dossier.Statut.REJETE, Dossier.Statut.NON_RETENU])
              .prefetch_related('historique')):
        h = next((x for x in d.historique.all() if x.nouveau_statut == d.statut), None)
        maj(d.texte_recherche, h.horodatage if h else d.modifie_le, 'dossier',
            h.motif if h else '', d.nom, d.postnom, d.prenom, d.email)

    for r in appel.reclamations.filter(statut=ReclamationEligibilite.Statut.REJETEE):
        maj(r.texte_recherche, r.traite_le, 'reclamation', r.motif, r.nom, r.postnom, r.prenom, r.email)

    for r in (Recours.objects.filter(statut=Recours.Statut.REJETE)
              .filter(Q(dossier__appel=appel) | Q(reclamation__appel=appel))):
        cle = normaliser_texte(f'{r.nom} {r.postnom} {r.prenom}')
        maj(cle, r.traite_le, 'recours', r.reponse, r.nom, r.postnom, r.prenom, r.email)

    non_retenus = [{'type': 'non_retenu', 'cle': k, 'nom': v['nom'], 'postnom': v['postnom'],
                    'prenom': v['prenom'], 'email': v['email'], 'motif': v['motif'],
                    'dernier_traitement': v['dernier_traitement'], 'date': v['date']}
                   for k, v in cands.items()]
    return list(admis.values()), non_retenus


class AppelCandidatureViewSet(viewsets.ModelViewSet):
    queryset = (
        AppelCandidature.objects
        .annotate(nb_dossiers=Count('dossiers'))
        .prefetch_related('pieces_exigees__type_piece')
    )
    serializer_class = AppelCandidatureSerializer
    # Consultable par tous (liste des appels, checklist des pièces) ; création
    # et modification réservées aux administrateurs. La config fine se fait
    # surtout en Django Admin.
    permission_classes = [EstAdminOuLectureSeule]

    # Publier/dépublier les retenus est une action de SUPERVISION (pas de la
    # config d'appel) : on lève la restriction admin-écriture du ViewSet et le
    # corps vérifie `peut_superviser`. Créer/éditer un appel reste admin.
    @action(detail=True, methods=['post'], url_path='publier-retenus',
            permission_classes=[IsAuthenticated])
    def publier_retenus(self, request, pk=None):
        """Publie la liste des retenus (admin/superviseur) → affichage public.

        C'est le SEUL moment où des emails partent : aucun email n'est envoyé
        pendant le traitement. On met en file (`EmailQueue`) l'email de
        convocation pour chaque retenu pas encore mis en file (idempotent :
        republier ne recrée pas de doublon). Les envois sont ensuite lissés par
        `python manage.py envoyer_emails_en_attente --limite N` (cron), pour
        respecter la limite quotidienne de Resend.
        """
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        appel = self.get_object()
        appel.liste_retenus_publiee = True
        appel.save(update_fields=['liste_retenus_publiee'])

        retenus = appel.dossiers.filter(statut=Dossier.Statut.RETENU)
        deja = set(
            EmailQueue.objects.filter(dossier__appel=appel)
            .values_list('dossier_id', flat=True)
        )
        def _code(d):
            return d.code or f'#{d.pk}'

        a_creer = [
            EmailQueue(
                dossier=d, destinataire=d.email,
                sujet=f'Résultat de votre candidature — dossier {_code(d)}',
                template='convocation_retenu.html',
                contexte={
                    'nom_candidat': f'{d.nom} {d.postnom} {d.prenom}'.strip(),
                    'code_dossier': _code(d),
                    'appel': appel.titre,
                },
            )
            for d in retenus if d.id not in deja and d.email
        ]
        EmailQueue.objects.bulk_create(a_creer)
        return Response({
            'detail': 'Liste des retenus publiée.',
            'retenus': retenus.count(),
            'emails_en_file': len(a_creer),
        })

    @action(detail=True, methods=['post'], url_path='depublier-retenus',
            permission_classes=[IsAuthenticated])
    def depublier_retenus(self, request, pk=None):
        """Retire la liste des retenus de l'affichage public (admin/superviseur)."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        appel = self.get_object()
        appel.liste_retenus_publiee = False
        appel.save(update_fields=['liste_retenus_publiee'])
        return Response({'detail': 'Liste des retenus dépubliée.'})

    # --- Liste DÉFINITIVE (retenus publiés + recours validés) -----------

    @action(detail=True, methods=['get'], url_path='liste-definitive',
            permission_classes=[IsAuthenticated])
    def liste_definitive(self, request, pk=None):
        """Aperçu (back-office) de la liste définitive : combinaison retenus +
        recours validés, avec le code déjà attribué (si publiée). Recherche `?q=`."""
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        appel = self.get_object()
        entrees = {e.texte_recherche: e for e in appel.retenus_definitifs.all()}
        rows = []
        for s in _sources_liste_definitive(appel):
            e = entrees.get(s['texte_recherche'])
            rows.append({
                **s,
                'id': e.id if e else None,
                'code': e.code if e else '',
                'ville_examen_code': e.ville_examen if e else '',
                'ville_examen': e.get_ville_examen_display() if e else '',
                'salle': e.salle if e else '',
                'ville_choisie': bool(e and e.ville_choisie_le) if e else False,
            })
        for token in tokens_recherche(request.query_params.get('q', '')):
            rows = [r for r in rows if token in r['texte_recherche']]
        # Tri par CODE croissant (codes sur 4 chiffres = ordre numérique) : les
        # ajouts supplémentaires (codes les plus élevés) restent en fin de liste.
        # Les entrées pas encore codées (avant publication) passent à la fin.
        rows.sort(key=lambda r: (r['code'] == '', r['code']))
        nb_recours = sum(1 for r in rows if r['origine'] == RetenuDefinitif.Origine.RECOURS)
        return Response({
            'publiee': appel.liste_definitive_publiee,
            'total': len(rows),
            'nb_recours': nb_recours,
            'nb_codes': len(entrees),
            'message': appel.message_retenus_definitif,
            'results': rows,
        })

    @action(detail=True, methods=['post'], url_path='publier-liste-definitive',
            permission_classes=[IsAuthenticated])
    def publier_liste_definitive(self, request, pk=None):
        """Publie la liste définitive (admin/superviseur) : fige les entrées,
        attribue les codes stables, et la rend publique (elle remplace alors la
        liste provisoire sur la page publique). Idempotent : republier n'ajoute
        que les nouvelles personnes (codes existants conservés)."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        appel = self.get_object()
        with transaction.atomic():
            crees = _generer_liste_definitive(appel)
            appel.liste_definitive_publiee = True
            appel.save(update_fields=['liste_definitive_publiee'])
        return Response({
            'detail': 'Liste définitive publiée.',
            'nouveaux_codes': crees,
            'total': appel.retenus_definitifs.count(),
        })

    @action(detail=True, methods=['post'], url_path='depublier-liste-definitive',
            permission_classes=[IsAuthenticated])
    def depublier_liste_definitive(self, request, pk=None):
        """Retire la liste définitive de l'affichage public (les codes déjà
        attribués sont conservés : ils restent stables si on republie)."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        appel = self.get_object()
        appel.liste_definitive_publiee = False
        appel.save(update_fields=['liste_definitive_publiee'])
        return Response({'detail': "Liste définitive retirée de l'affichage public."})

    # --- Liste des admis à l'INTERVIEW (étape suivant le test) ----------

    @action(detail=True, methods=['post'], url_path='publier-liste-interview',
            permission_classes=[IsAuthenticated])
    def publier_liste_interview(self, request, pk=None):
        """Publie la liste des admis à l'interview (admin/superviseur). La page
        publique n'affiche alors plus que ce sous-ensemble (avec son communiqué).
        Les admis à l'interview doivent d'abord être marqués (import)."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        appel = self.get_object()
        nb = appel.retenus_definitifs.filter(admis_interview=True).count()
        if not nb:
            raise ValidationError("Aucun admis à l'interview n'est marqué pour cet appel "
                                  "(importez d'abord la liste des admis à l'interview).")
        appel.liste_interview_publiee = True
        appel.save(update_fields=['liste_interview_publiee'])
        return Response({'detail': "Liste des admis à l'interview publiée.", 'total': nb})

    @action(detail=True, methods=['post'], url_path='depublier-liste-interview',
            permission_classes=[IsAuthenticated])
    def depublier_liste_interview(self, request, pk=None):
        """Retire la liste des admis à l'interview de l'affichage public : la page
        publique réaffiche la liste définitive (complète)."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        appel = self.get_object()
        appel.liste_interview_publiee = False
        appel.save(update_fields=['liste_interview_publiee'])
        return Response({'detail': "Liste des admis à l'interview retirée de l'affichage public."})

    # --- Demandes de ville d'examen (validées par un agent) -------------

    @action(detail=True, methods=['get'], url_path='demandes-ville',
            permission_classes=[IsAuthenticated])
    def demandes_ville(self, request, pk=None):
        """Demandes de ville EN ATTENTE pour cet appel (back-office)."""
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        appel = self.get_object()
        S = RetenuDefinitif.DemandeStatut
        base = appel.retenus_definitifs.exclude(ville_demande_statut='')
        counts = {
            'en_attente': base.filter(ville_demande_statut=S.EN_ATTENTE).count(),
            'validees': base.filter(ville_demande_statut=S.VALIDEE).count(),
            'rejetees': base.filter(ville_demande_statut=S.REJETEE).count(),
        }
        statut = self.request.query_params.get('statut') or S.EN_ATTENTE
        qs = base
        if statut in (S.EN_ATTENTE, S.VALIDEE, S.REJETEE):
            qs = base.filter(ville_demande_statut=statut)
        qs = (qs.select_related('dossier', 'recours__dossier', 'recours__reclamation',
                                'ville_traite_par')
              .order_by('ville_demandee_le'))
        rows = [{
            'id': e.id, 'code': e.code,
            'nom': e.nom, 'postnom': e.postnom, 'prenom': e.prenom,
            'poste_libelle': e.poste_libelle,
            'ville_actuelle': e.get_ville_examen_display(),
            'ville_demandee': e.ville_demandee,
            'ville_demandee_libelle': e.get_ville_demandee_display(),
            'statut': e.ville_demande_statut,
            'traite_par': (e.ville_traite_par.get_full_name() or e.ville_traite_par.email) if e.ville_traite_par else '',
            'date_naissance': e.date_naissance,
            'demande_le': e.ville_demandee_le,
            'documents': _docs_retenu_definitif(e),
        } for e in qs]
        return Response({**counts, 'statut': statut, 'total': len(rows), 'results': rows})

    @action(detail=True, methods=['post'], url_path='traiter-demande-ville',
            permission_classes=[IsAuthenticated])
    def traiter_demande_ville(self, request, pk=None):
        """Valide ou rejette une demande de ville (back-office). Corps :
        { id, action: 'valider'|'rejeter' }. À la validation, la ville demandée
        devient la ville officielle ; au rejet, la demande est annulée (la ville
        reste inchangée)."""
        if not roles.peut_traiter(request.user):
            raise PermissionDenied("Réservé aux administrateurs, superviseurs et validateurs.")
        appel = self.get_object()
        entree = get_object_or_404(
            RetenuDefinitif, pk=request.data.get('id'), appel=appel,
        )
        if entree.ville_demande_statut != RetenuDefinitif.DemandeStatut.EN_ATTENTE:
            raise ValidationError("Aucune demande de ville en attente pour ce candidat.")
        action_ = request.data.get('action')
        # On CONSERVE `ville_demandee` (historique) ; seul le statut change.
        if action_ == 'valider':
            entree.ville_examen = entree.ville_demandee
            entree.ville_choisie_le = timezone.now()
            entree.ville_traite_par = request.user
            entree.ville_demande_statut = RetenuDefinitif.DemandeStatut.VALIDEE
            entree.save(update_fields=['ville_examen', 'ville_choisie_le',
                                       'ville_traite_par', 'ville_demande_statut'])
        elif action_ == 'rejeter':
            entree.ville_traite_par = request.user
            entree.ville_demande_statut = RetenuDefinitif.DemandeStatut.REJETEE
            entree.save(update_fields=['ville_traite_par', 'ville_demande_statut'])
        else:
            raise ValidationError({'action': "Action invalide (valider/rejeter)."})
        return Response({
            'detail': 'Demande validée.' if action_ == 'valider' else 'Demande rejetée.',
            'ville_examen': entree.ville_examen,
            'ville_examen_libelle': entree.get_ville_examen_display(),
        })

    @action(detail=True, methods=['post'], url_path='affecter-salles',
            permission_classes=[IsAuthenticated])
    def affecter_salles(self, request, pk=None):
        """Affecte automatiquement les salles d'une ville : les candidats de cette
        ville (triés par CODE) sont répartis par paquets de `par_salle` dans des
        salles A, B, C… (jusqu'à `nombre_salles`). Ré-exécutable : les codes étant
        stables, chacun retrouve la même salle."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        appel = self.get_object()
        ville = request.data.get('ville')
        if ville not in RetenuDefinitif.Ville.values:
            raise ValidationError({'ville': 'Ville invalide.'})
        try:
            nb_salles = int(request.data.get('nombre_salles'))
            par_salle = int(request.data.get('par_salle'))
        except (TypeError, ValueError):
            raise ValidationError("Indiquez le nombre de salles et le nombre de personnes par salle.")
        if nb_salles < 1 or par_salle < 1:
            raise ValidationError("Le nombre de salles et de personnes par salle doit être ≥ 1.")

        from collections import defaultdict
        entries = appel.retenus_definitifs.filter(ville_examen=ville).order_by('code')
        # Si les suppléments sont masqués, ils ne reçoivent pas de salle.
        if not appel.afficher_supplements_definitif:
            entries = entries.exclude(origine=RetenuDefinitif.Origine.SUPPLEMENT)
        ids = list(entries.values_list('id', flat=True))
        par_room = defaultdict(list)
        non_affectes = 0
        for i, id_ in enumerate(ids):
            ri = i // par_salle
            if ri >= nb_salles:
                non_affectes += 1
                continue
            par_room[ri].append(id_)
        with transaction.atomic():
            entries.update(salle='')   # réinitialise la ville
            for ri, group in par_room.items():
                RetenuDefinitif.objects.filter(id__in=group).update(salle=_lettre_salle(ri))
        return Response({
            'ville': ville,
            'total': len(ids),
            'affectes': len(ids) - non_affectes,
            'salles_utilisees': len(par_room),
            'par_salle': par_salle,
            'non_affectes': non_affectes,
        })

    @action(detail=True, methods=['post'], url_path='afficher-salle-public',
            permission_classes=[IsAuthenticated])
    def afficher_salle_public(self, request, pk=None):
        """Active/désactive l'affichage de la salle sur la page publique (badge
        + liste définitive). Corps : { afficher: bool } (superviseur)."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        appel = self.get_object()
        appel.afficher_salle_public = bool(request.data.get('afficher'))
        appel.save(update_fields=['afficher_salle_public'])
        return Response({'afficher_salle_public': appel.afficher_salle_public})

    @action(detail=True, methods=['post'], url_path='afficher-supplements',
            permission_classes=[IsAuthenticated])
    def afficher_supplements(self, request, pk=None):
        """Active/désactive l'inclusion des ajouts SUPPLÉMENTAIRES dans la liste
        définitive (page publique, aperçu, PDF, Excel, salles). Corps :
        { afficher: bool } (superviseur). Ne supprime pas les entrées."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        appel = self.get_object()
        appel.afficher_supplements_definitif = bool(request.data.get('afficher'))
        appel.save(update_fields=['afficher_supplements_definitif'])
        return Response({'afficher_supplements_definitif': appel.afficher_supplements_definitif})

    @action(detail=True, methods=['get'], url_path='liste-definitive-export',
            permission_classes=[IsAuthenticated])
    def liste_definitive_export(self, request, pk=None):
        """Exporte la liste définitive en Excel (back-office) : code, identité,
        domaine, ville du test, demande de ville en attente, origine."""
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        appel = self.get_object()

        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Liste définitive'
        entetes = ['Code', 'Nom', 'Postnom', 'Prénom', 'Domaine', 'Ville du test',
                   'Salle', 'Demande en attente', 'Origine']
        ws.append(entetes)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        export_qs = appel.retenus_definitifs.all().order_by('code')
        if not appel.afficher_supplements_definitif:
            export_qs = export_qs.exclude(origine=RetenuDefinitif.Origine.SUPPLEMENT)
        for e in export_qs:
            ws.append([
                e.code, e.nom, e.postnom, e.prenom, e.poste_libelle,
                e.get_ville_examen_display(), e.salle,
                e.get_ville_demandee_display() if e.ville_demandee else '',
                e.get_origine_display(),
            ])
        for i, largeur in enumerate([10, 20, 20, 20, 28, 16, 8, 18, 22], start=1):
            ws.column_dimensions[get_column_letter(i)].width = largeur
        # En-tête figé + tri/filtre Excel activé sur les colonnes.
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        reponse = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        reponse['Content-Disposition'] = 'attachment; filename="liste_definitive.xlsx"'
        wb.save(reponse)
        return reponse

    # --- E-mails de résultat (aperçu / export ; envoi en phase 2) --------

    @action(detail=True, methods=['get'], url_path='resultats-apercu',
            permission_classes=[IsAuthenticated])
    def resultats_apercu(self, request, pk=None):
        """APERÇU (aucun envoi) des e-mails de résultat : 1 par personne, admis
        (liste définitive) + non retenus (dernier traitement). Réservé admin."""
        if not roles.est_admin(request.user):
            raise PermissionDenied("Réservé aux administrateurs.")
        appel = self.get_object()
        admis, non_retenus = _resultats_emails(appel)
        rows = admis + non_retenus
        sans_email = sum(1 for r in rows if not r['email'])
        return Response({
            'admis': len(admis),
            'non_retenus': len(non_retenus),
            'total': len(rows),
            'sans_email': sans_email,
            'results': rows,
        })

    @action(detail=True, methods=['get'], url_path='resultats-export',
            permission_classes=[IsAuthenticated])
    def resultats_export(self, request, pk=None):
        """Export Excel de l'aperçu des e-mails de résultat (admin)."""
        if not roles.est_admin(request.user):
            raise PermissionDenied("Réservé aux administrateurs.")
        appel = self.get_object()
        admis, non_retenus = _resultats_emails(appel)

        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Résultats e-mail'
        ws.append(['Résultat', 'Code', 'Nom', 'Postnom', 'Prénom', 'E-mail',
                   'Dernier traitement', 'Ville / Motif'])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in admis:
            ws.append(['ADMIS', r['code'], r['nom'], r['postnom'], r['prenom'],
                       r['email'], 'liste définitive', r['ville']])
        for r in non_retenus:
            ws.append(['NON RETENU', '', r['nom'], r['postnom'], r['prenom'],
                       r['email'], r['dernier_traitement'], r['motif']])
        for i, largeur in enumerate([13, 10, 20, 20, 20, 28, 18, 40], start=1):
            ws.column_dimensions[get_column_letter(i)].width = largeur
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        reponse = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        reponse['Content-Disposition'] = 'attachment; filename="resultats_emails.xlsx"'
        wb.save(reponse)
        return reponse

    # --- Envoi des e-mails de résultat (file EmailQueue, progression) ----

    SUJET_RESULTAT = 'Résultat de votre candidature — ACGT'

    @staticmethod
    def _email_resultat(row):
        """(template, contexte) pour une ligne consolidée (admis / non retenu)."""
        nom = ' '.join(x for x in [row['nom'], row.get('postnom', ''), row['prenom']] if x).strip()
        if row['type'] == 'admis':
            return 'resultat_admis.html', {'nom': nom, 'code': row.get('code', ''), 'ville': row.get('ville', '')}
        return 'resultat_non_retenu.html', {'nom': nom, 'motif': row.get('motif', '')}

    def _etat_resultats(self, appel):
        qs = EmailQueue.objects.filter(appel=appel, categorie='resultat')
        total = qs.count()
        envoyes = qs.filter(statut=EmailQueue.Statut.ENVOYE).count()
        echecs = qs.filter(statut=EmailQueue.Statut.ECHEC).count()
        restants = qs.filter(statut=EmailQueue.Statut.EN_ATTENTE).count()
        return {'total': total, 'envoyes': envoyes, 'echecs': echecs,
                'restants': restants, 'termine': restants == 0}

    @action(detail=True, methods=['get'], url_path='resultats-etat',
            permission_classes=[IsAuthenticated])
    def resultats_etat(self, request, pk=None):
        """État de la campagne d'envoi des résultats (progression)."""
        if not roles.est_admin(request.user):
            raise PermissionDenied("Réservé aux administrateurs.")
        return Response(self._etat_resultats(self.get_object()))

    @action(detail=True, methods=['post'], url_path='resultats-test',
            permission_classes=[IsAuthenticated])
    def resultats_test(self, request, pk=None):
        """Envoie les 2 exemples (admis + non retenu) à l'adresse de l'admin."""
        if not roles.est_admin(request.user):
            raise PermissionDenied("Réservé aux administrateurs.")
        appel = self.get_object()
        admis, non = _resultats_emails(appel)
        ex_a = admis[0] if admis else {'type': 'admis', 'nom': 'NOM POSTNOM PRÉNOM', 'postnom': '', 'prenom': '', 'code': '0001', 'ville': 'Kinshasa'}
        ex_n = non[0] if non else {'type': 'non_retenu', 'nom': 'NOM POSTNOM PRÉNOM', 'postnom': '', 'prenom': '', 'motif': 'Critères non remplis'}
        dest = request.user.email
        try:
            for ex in (ex_a, ex_n):
                tmpl, ctx = self._email_resultat(ex)
                envoyer_email(dest, '[TEST] ' + self.SUJET_RESULTAT, tmpl, ctx)
        except EmailError:
            return Response({'detail': "Échec de l'envoi de test (service e-mail)."}, status=502)
        return Response({'detail': f'2 exemples envoyés à {dest}.'})

    @action(detail=True, methods=['post'], url_path='resultats-preparer',
            permission_classes=[IsAuthenticated])
    def resultats_preparer(self, request, pk=None):
        """Met en file les e-mails de résultat (idempotent, 1 par adresse).
        N'envoie rien : prépare la file que `resultats-envoyer-lot` videra."""
        if not roles.est_admin(request.user):
            raise PermissionDenied("Réservé aux administrateurs.")
        appel = self.get_object()
        admis, non = _resultats_emails(appel)
        existants = set(
            EmailQueue.objects.filter(appel=appel, categorie='resultat')
            .values_list('destinataire', flat=True)
        )
        a_creer, vus = [], set()
        prepares = deja = sans_email = 0
        for r in admis + non:
            em = (r['email'] or '').strip().lower()
            if not em:
                sans_email += 1
                continue
            if em in existants or em in vus:
                deja += 1
                continue
            vus.add(em)
            tmpl, ctx = self._email_resultat(r)
            a_creer.append(EmailQueue(
                appel=appel, categorie='resultat', destinataire=em,
                sujet=self.SUJET_RESULTAT, template=tmpl, contexte=ctx,
                statut=EmailQueue.Statut.EN_ATTENTE,
            ))
            prepares += 1
        EmailQueue.objects.bulk_create(a_creer)
        etat = self._etat_resultats(appel)
        return Response({'prepares': prepares, 'deja_en_file': deja,
                         'sans_email': sans_email, **etat})

    @action(detail=True, methods=['post'], url_path='resultats-envoyer-lot',
            permission_classes=[IsAuthenticated])
    def resultats_envoyer_lot(self, request, pk=None):
        """Envoie le prochain lot d'e-mails de résultat en attente (progression).
        Appelé en boucle par le front jusqu'à `termine`."""
        if not roles.est_admin(request.user):
            raise PermissionDenied("Réservé aux administrateurs.")
        appel = self.get_object()
        limite = max(1, min(int(request.data.get('limite') or 30), 100))
        lot = list(
            EmailQueue.objects.filter(appel=appel, categorie='resultat',
                                      statut=EmailQueue.Statut.EN_ATTENTE)
            .order_by('cree_le')[:limite]
        )
        for e in lot:
            try:
                envoyer_email(e.destinataire, e.sujet, e.template, e.contexte)
            except EmailError:
                e.tentatives += 1
                if e.tentatives >= EmailQueue.MAX_TENTATIVES:
                    e.statut = EmailQueue.Statut.ECHEC
                e.save(update_fields=['tentatives', 'statut'])
                continue
            e.statut = EmailQueue.Statut.ENVOYE
            e.envoye_le = timezone.now()
            e.save(update_fields=['statut', 'envoye_le'])
        return Response(self._etat_resultats(appel))


class RetenusViewSet(viewsets.ReadOnlyModelViewSet):
    """Liste publique des personnes retenues (lecture seule, recherche tolérante).

    N'expose que les dossiers RETENU d'un AAC dont la liste est publiée, en
    NOM/POSTNOM/PRÉNOM uniquement. Filtrer par `?appel=<id>` et rechercher
    via `?q=`.
    """

    serializer_class = RetenuPubliqueSerializer
    permission_classes = [AllowAny]
    pagination_class = PaginationPublique

    def get_queryset(self):
        qs = Dossier.objects.filter(
            statut=Dossier.Statut.RETENU,
            appel__liste_retenus_publiee=True,
        ).select_related('poste')
        appel = self.request.query_params.get('appel')
        if appel:
            qs = qs.filter(appel_id=appel)
        for token in tokens_recherche(self.request.query_params.get('q', '')):
            qs = qs.filter(texte_recherche__contains=token)
        return qs.order_by('nom', 'postnom', 'prenom')


class VilleExamenThrottle(AnonRateThrottle):
    """Limite le formulaire public de choix de ville (anti-spam)."""

    scope = 'reclamation'


class RetenuSupplementViewSet(viewsets.ModelViewSet):
    """CRUD des ajouts SUPPLÉMENTAIRES à la liste définitive (origine=supplement) :
    back-office, réservé aux administrateurs. Le CODE est attribué automatiquement
    à la suite du dernier code définitif de l'appel (stable). On ne touche jamais
    les entrées normales (liste/recours) : ce ViewSet ne voit que les suppléments."""

    serializer_class = RetenuSupplementSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        if not roles.acces_backoffice(self.request.user):
            return RetenuDefinitif.objects.none()
        qs = RetenuDefinitif.objects.filter(origine=RetenuDefinitif.Origine.SUPPLEMENT)
        appel = self.request.query_params.get('appel')
        if appel:
            qs = qs.filter(appel_id=appel)
        return qs.order_by('code')

    def _verifier_admin(self):
        if not roles.est_admin(self.request.user):
            raise PermissionDenied("La gestion des ajouts supplémentaires est réservée aux administrateurs.")

    def perform_create(self, serializer):
        self._verifier_admin()
        appel = serializer.validated_data['appel']
        codes = [int(e.code) for e in appel.retenus_definitifs.all() if e.code.isdigit()]
        prochain = (max(codes) + 1) if codes else 1
        serializer.save(code=f'{prochain:04d}', origine=RetenuDefinitif.Origine.SUPPLEMENT)

    def perform_update(self, serializer):
        self._verifier_admin()
        serializer.save()

    def perform_destroy(self, instance):
        self._verifier_admin()
        instance.delete()


class RetenusDefinitifsViewSet(viewsets.ReadOnlyModelViewSet):
    """Liste publique DÉFINITIVE (lecture seule) : CODE stable + identité +
    domaine, pour les appels dont la liste définitive est publiée. `?appel=`/`?q=`.

    Action publique `choisir-ville` : un retenu hors Kinshasa précise la ville
    du test (Lubumbashi / Mbuji-Mayi). Kinshasa reste la valeur par défaut."""

    serializer_class = RetenuDefinitifSerializer
    permission_classes = [AllowAny]
    pagination_class = PaginationPublique

    def get_throttles(self):
        if self.action == 'choisir_ville':
            return [VilleExamenThrottle()]
        return super().get_throttles()

    def _base_public(self):
        """Sous-ensemble public de base (avant recherche/filtre domaine) :
        listes publiées, suppléments éventuellement masqués, et — si l'interview
        est publiée — uniquement les admis à l'interview."""
        qs = RetenuDefinitif.objects.filter(appel__liste_definitive_publiee=True)
        qs = qs.exclude(origine=RetenuDefinitif.Origine.SUPPLEMENT,
                        appel__afficher_supplements_definitif=False)
        qs = qs.exclude(appel__liste_interview_publiee=True, admis_interview=False)
        appel = self.request.query_params.get('appel')
        if appel:
            qs = qs.filter(appel_id=appel)
        return qs

    def get_queryset(self):
        qs = self._base_public()
        # Filtre par domaine (sélection sur la page publique).
        domaine = (self.request.query_params.get('domaine') or '').strip()
        if domaine:
            qs = qs.filter(poste_libelle=domaine)
        # Recherche par CODE (ex. « 0001 » ou « 1 ») : utilisée par le formulaire
        # de choix de ville (le candidat saisit son code et retrouve son nom).
        code = (self.request.query_params.get('code') or '').strip()
        if code:
            qs = qs.filter(code=code.zfill(4) if code.isdigit() else code)
        for token in tokens_recherche(self.request.query_params.get('q', '')):
            qs = qs.filter(texte_recherche__contains=token)
        # En mode INTERVIEW, on respecte l'ORDRE DU FICHIER (interview_ordre) ;
        # sinon (liste définitive/test), l'ordre reste celui des codes.
        if self._mode_interview(self.request.query_params.get('appel')):
            return qs.order_by('interview_ordre', 'code')
        return qs.order_by('code')

    @action(detail=False, methods=['get'], url_path='domaines',
            permission_classes=[AllowAny])
    def domaines(self, request):
        """Liste des domaines présents dans la liste publiée (+ effectifs), pour
        alimenter le sélecteur de domaine de la page publique. En mode interview,
        les domaines suivent l'ordre du fichier ; sinon, ordre alphabétique."""
        from django.db.models import Count, Min
        qs = self._base_public()
        if self._mode_interview(request.query_params.get('appel')):
            rows = (qs.values('poste_libelle')
                      .annotate(n=Count('id'), o=Min('interview_ordre')).order_by('o'))
        else:
            rows = (qs.values('poste_libelle')
                      .annotate(n=Count('id')).order_by('poste_libelle'))
        return Response([{'domaine': r['poste_libelle'], 'count': r['n']}
                         for r in rows if (r['poste_libelle'] or '').strip()])

    def _mode_interview(self, appel_id):
        """Vrai si la liste des admis à l'interview est publiée (pour l'appel
        demandé, sinon pour au moins un appel publié)."""
        aq = AppelCandidature.objects.filter(liste_definitive_publiee=True,
                                             liste_interview_publiee=True)
        if appel_id:
            aq = aq.filter(id=appel_id)
        return aq.exists()

    @action(detail=False, methods=['post'], url_path='choisir-ville',
            permission_classes=[AllowAny])
    def choisir_ville(self, request):
        """Le retenu DEMANDE sa ville de test (public). Corps : { id,
        date_naissance, ville }. `ville` ∈ {lubumbashi, mbuji_mayi}.

        SÉCURITÉ : on n'écrit PAS la ville officielle ici (un formulaire public
        ne doit pas pouvoir modifier directement la ville de n'importe qui). On
        enregistre une DEMANDE en attente, qu'un agent valide en back-office."""
        entree = get_object_or_404(
            RetenuDefinitif, pk=request.data.get('id'),
            appel__liste_definitive_publiee=True,
        )
        ville = request.data.get('ville')
        # On n'autorise que les villes hors Kinshasa (Kinshasa = défaut, pas de demande).
        if ville not in (RetenuDefinitif.Ville.LUBUMBASHI, RetenuDefinitif.Ville.MBUJI_MAYI):
            raise ValidationError({'ville': "Choisissez Lubumbashi ou Mbuji-Mayi."})
        dn = request.data.get('date_naissance')
        if not dn:
            raise ValidationError({'date_naissance': "La date de naissance est requise."})
        entree.ville_demandee = ville
        entree.date_naissance = dn
        entree.ville_demandee_le = timezone.now()
        entree.ville_demande_statut = RetenuDefinitif.DemandeStatut.EN_ATTENTE
        # Nouvelle demande : on repart d'une décision vierge.
        entree.ville_choisie_le = None
        entree.ville_traite_par = None
        entree.save(update_fields=['ville_demandee', 'date_naissance', 'ville_demandee_le',
                                   'ville_demande_statut', 'ville_choisie_le', 'ville_traite_par'])
        return Response({
            'detail': 'Demande enregistrée (en attente de validation).',
            'ville': entree.ville_demandee,
            'ville_libelle': entree.get_ville_demandee_display(),
        })


class DossierViewSet(viewsets.ModelViewSet):
    serializer_class = DossierSerializer
    pagination_class = PaginationStandard

    # Champs autorisés au tri (allowlist : évite l'injection de champs arbitraires).
    TRI_AUTORISE = {'id', 'code', 'nom', 'postnom', 'prenom', 'statut', 'cree_le',
                    'poste__libelle', 'appel__titre'}

    def get_serializer_class(self):
        # Liste = vue allégée (pas de N+1 sur pièces/complétude) ; détail et
        # réponses d'action = vue complète.
        if self.action == 'list':
            return DossierListeSerializer
        return DossierSerializer

    @staticmethod
    def _annoter_correspondance(qs):
        """Ajoute les annotations de correspondance avec la liste d'éligibilité.

        Réutilisé par `get_queryset` (badges) et par `stats_correspondance`
        (histogramme) pour garantir des définitions identiques. Comparaison
        insensible à la casse ; un champ ne compte que renseigné des deux côtés.
        """
        return qs.annotate(
            corresp_code=Exists(
                ListeEligibilite.objects.exclude(code='')
                .filter(code__iexact=OuterRef('code'))
            ),
            # Nom complet (nom+postnom+prénom) trouvé sur une même ligne →
            # « à rattacher » : c'est très probablement la personne.
            corresp_nom_complet=Exists(
                ListeEligibilite.objects.exclude(texte_recherche='')
                .filter(texte_recherche=OuterRef('texte_recherche'))
            ),
            corresp_f_nom=Exists(
                ListeEligibilite.objects.exclude(nom='')
                .filter(nom__iexact=OuterRef('nom'))
            ),
            corresp_f_postnom=Exists(
                ListeEligibilite.objects.exclude(postnom='')
                .filter(postnom__iexact=OuterRef('postnom'))
            ),
            corresp_f_prenom=Exists(
                ListeEligibilite.objects.exclude(prenom='')
                .filter(prenom__iexact=OuterRef('prenom'))
            ),
        )

    @staticmethod
    def _doublon_exists():
        """Sous-requête : ce dossier a-t-il un jumeau SOUMIS (hors brouillon et
        rejeté) du même appel et même nom complet ? Réutilisée par l'annotation
        de liste et par la navigation « doublon suivant »."""
        return Exists(
            Dossier.objects
            .filter(appel_id=OuterRef('appel_id'))
            .exclude(statut__in=[Dossier.Statut.BROUILLON, Dossier.Statut.REJETE])
            .exclude(texte_recherche='')
            .filter(texte_recherche=OuterRef('texte_recherche'))
            .exclude(pk=OuterRef('pk'))
        )

    def _filtrer_liste(self, qs, params):
        """Filtres de la liste sur un queryset DÉJÀ annoté (corresp_* + a_doublon).

        Partagé par `get_queryset` (affichage) et `repartir` (répartition du
        MÊME sous-ensemble) → les buckets de correspondance restent identiques.
        `params` accepte un QueryDict (GET) comme un dict (corps JSON) : `.get`
        fonctionne dans les deux cas.
        """
        statut = params.get('statut')
        appel = params.get('appel')
        if statut:
            # Statut composite possible (« depose,en_examen » = À valider).
            statuts = [s for s in str(statut).split(',') if s]
            qs = qs.filter(statut__in=statuts) if len(statuts) > 1 else qs.filter(statut=statuts[0])
        if appel:
            qs = qs.filter(appel_id=appel)

        # Correspondance avec la liste d'éligibilité (alignée sur les badges).
        corr = params.get('correspondance')
        if corr == 'rattache':
            qs = qs.filter(ligne_eligibilite__isnull=False)
        elif corr == 'a_rattacher':
            qs = qs.filter(ligne_eligibilite__isnull=True, corresp_nom_complet=True)
        elif corr == 'partielle':
            # Au moins un champ de NOM coïncide (ni rattaché ni nom complet). Le
            # code seul ne compte pas (il peut être celui d'autrui) → « aucune ».
            qs = qs.filter(
                ligne_eligibilite__isnull=True, corresp_nom_complet=False,
            ).filter(
                Q(corresp_f_nom=True) | Q(corresp_f_postnom=True)
                | Q(corresp_f_prenom=True)
            )
        elif corr == 'aucune':
            # Aucun champ de NOM ne coïncide (le code éventuel est ignoré).
            qs = qs.filter(
                ligne_eligibilite__isnull=True, corresp_nom_complet=False,
                corresp_f_nom=False, corresp_f_postnom=False, corresp_f_prenom=False,
            )

        # Doublons uniquement : dossiers ayant au moins un jumeau.
        if params.get('doublons') in ('1', 'true', 'oui', True, 1):
            qs = qs.filter(a_doublon=True)

        # Recherche tolérante par nom (accents/casse/ordre) OU par code.
        q = (params.get('q') or '').strip()
        if q:
            tokens = tokens_recherche(q)
            if tokens:
                nom_q = Q()
                for token in tokens:
                    nom_q &= Q(texte_recherche__contains=token)
                qs = qs.filter(nom_q | Q(code__icontains=q))
            else:
                qs = qs.filter(code__icontains=q)
        return qs

    def get_queryset(self):
        """Scoping par rôle :

        - admin : tous les dossiers ;
        - évaluateur : uniquement les dossiers où il est désigné (+ les siens
          s'il a aussi déposé) ;
        - candidat : uniquement ses propres dossiers.
        """
        # Correspondance avec la liste d'éligibilité (badge indicatif, jamais
        # bloquant) : on indique précisément quels champs coïncident (code, nom,
        # postnom, prénom). Comparaison insensible à la casse (`iexact`) ;
        # l'insensibilité aux accents viendra avec `unaccent` en prod (cf.
        # CLAUDE.md). Un champ ne compte que s'il est renseigné des deux côtés
        # (jamais de faux match sur un champ vide). Calcul en SQL (Exists) pour
        # éviter tout N+1.
        qs = self._annoter_correspondance(
            Dossier.objects
            .select_related('appel', 'deposant', 'ligne_eligibilite', 'affecte_a')
            .prefetch_related('pieces__type_piece')
        ).annotate(
            # Doublon probable : un AUTRE dossier SOUMIS (hors brouillon) du
            # même appel a le même NOM COMPLET (nom+postnom+prénom normalisé).
            # On n'utilise PAS l'email : un proche peut déposer plusieurs
            # dossiers (personnes différentes) depuis la même adresse. On
            # ignore les brouillons (non traités). Indicatif : l'admin tranche.
            a_doublon=self._doublon_exists(),
            # Origine : True si le dossier a été créé en validant une réclamation.
            est_reclamation=Exists(
                ReclamationEligibilite.objects.filter(dossier_cree=OuterRef('pk'))
            ),
        )
        user = self.request.user
        if roles.acces_backoffice(user):
            pass  # tout rôle back-office (admin, validateur, correcteur, lecteur) voit tout
        elif roles.est_evaluateur(user):
            qs = qs.filter(
                Q(affectations__evaluateur=user) | Q(deposant=user)
            ).distinct()
        else:
            qs = qs.filter(deposant=user)

        # Filtre d'affectation : « moi » = mon lot, « aucune » = non affectés,
        # un id = le lot d'un agent (pour l'admin qui supervise la répartition).
        affecte = self.request.query_params.get('affecte')
        if affecte == 'moi':
            qs = qs.filter(affecte_a=user)
        elif affecte == 'aucune':
            qs = qs.filter(affecte_a__isnull=True)
        elif affecte and affecte.isdigit():
            qs = qs.filter(affecte_a_id=int(affecte))

        # Filtres de liste (statut, appel, correspondance, doublons, recherche),
        # factorisés pour être réutilisés à l'identique par `repartir`.
        qs = self._filtrer_liste(qs, self.request.query_params)

        # Filtre d'origine : réclamation (validée depuis une réclamation) vs
        # candidature en ligne (déposée par un compte candidat).
        origine = self.request.query_params.get('origine')
        if origine == 'reclamation':
            qs = qs.filter(est_reclamation=True)
        elif origine == 'en_ligne':
            qs = qs.filter(est_reclamation=False)

        # Exclure les résidus annulés (dossiers de validation orphelins :
        # deposant=None + non liés à une réclamation) — vue « dossiers reçus réels ».
        if self.request.query_params.get('hors_residus') in ('1', 'true'):
            qs = qs.exclude(deposant__isnull=True, est_reclamation=False)

        # Tri demandé par le tableau (sinon : plus récents d'abord).
        ordering = self.request.query_params.get('ordering', '')
        if ordering.lstrip('-') in self.TRI_AUTORISE:
            qs = qs.order_by(ordering)
        else:
            qs = qs.order_by('-cree_le')
        return qs

    # --- Cycle de vie du dépôt (candidat) -------------------------------

    def perform_create(self, serializer):
        """Création d'un dossier en brouillon par un candidat à l'email vérifié."""
        user = self.request.user
        if not user.email_verifie:
            raise PermissionDenied(
                "Vous devez d'abord vérifier votre email avant de déposer un dossier."
            )
        appel = serializer.validated_data['appel']
        # Candidatures ouvertes uniquement : pas de dépôt sur un appel clôturé
        # (ou non publié). Garde-fou serveur, en plus du masquage côté front.
        if not appel.est_ouvert:
            raise ValidationError(
                "Les candidatures pour cet appel sont clôturées."
            )
        # Appel à candidature unique : un seul dossier par compte.
        if appel.candidature_unique and appel.dossiers.filter(deposant=user).exists():
            raise ValidationError(
                "Vous avez déjà une candidature pour cet appel : une seule est autorisée."
            )
        # Anti-doublon : un seul brouillon à la fois par appel et par compte.
        # Règle applicative à la création uniquement (pas de contrainte en
        # base) : les brouillons multiples existants restent valides.
        if appel.dossiers.filter(
            deposant=user, statut=Dossier.Statut.BROUILLON,
        ).exists():
            raise ValidationError(
                "Vous avez déjà un dossier en brouillon pour cet appel. "
                "Reprenez-le depuis « Mes dossiers » (ou supprimez-le) "
                "avant d'en créer un nouveau."
            )
        serializer.save(deposant=user, statut=Dossier.Statut.BROUILLON)

    def _verifier_modifiable(self, dossier):
        """Un dossier n'est éditable (CRUD, pièces) qu'en brouillon par son déposant."""
        if dossier.deposant_id != self.request.user.id and not self.request.user.is_superuser:
            raise PermissionDenied("Ce dossier ne vous appartient pas.")
        if not dossier.modifiable:
            raise ValidationError(
                "Ce dossier a déjà été soumis et n'est plus modifiable."
            )

    def update(self, request, *args, **kwargs):
        self._verifier_modifiable(self.get_object())
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        self._verifier_modifiable(self.get_object())
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['patch'], url_path='identite')
    def modifier_identite(self, request, pk=None):
        """Corrige l'identité d'un dossier : code, nom, postnom, prénom.

        Réservé aux administrateurs et correcteurs ; possible quel que soit le
        statut (sert à corriger une coquille après dépôt). `texte_recherche` est
        recalculé automatiquement (recherche/doublons cohérents).
        """
        if not (roles.est_admin(request.user) or roles.est_correcteur(request.user)):
            raise PermissionDenied(
                "Seuls les administrateurs et correcteurs peuvent modifier l'identité."
            )
        dossier = self.get_object()
        serializer = ModificationIdentiteSerializer(dossier, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(self.get_serializer(dossier).data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Comptes de dossiers par statut (scopés par rôle), pour les KPI.

        Reprend le scoping de `get_queryset` mais sans le filtre `statut`, afin
        de fournir les totaux de chaque statut en une seule requête.
        """
        user = request.user
        qs = Dossier.objects.all()
        if roles.acces_backoffice(user):
            pass
        elif roles.est_evaluateur(user):
            qs = qs.filter(Q(affectations__evaluateur=user) | Q(deposant=user)).distinct()
        else:
            qs = qs.filter(deposant=user)

        appel = request.query_params.get('appel')
        if appel:
            qs = qs.filter(appel_id=appel)

        par_statut = {row['statut']: row['n'] for row in qs.values('statut').annotate(n=Count('id'))}
        # Origine des RETENUS (validés) : issus d'une réclamation vs déposés en
        # ligne (liste des éligibles). Comptes sur le statut RETENU uniquement.
        retenus = qs.filter(statut=Dossier.Statut.RETENU)
        nb_reclam_val = retenus.filter(
            Exists(ReclamationEligibilite.objects.filter(dossier_cree=OuterRef('pk')))
        ).count()
        nb_retenu = retenus.count()
        total = sum(par_statut.values())
        # Reçus « réels » = hors brouillon ET hors résidus annulés (dossiers
        # créés en validant une réclamation puis annulés : deposant=None +
        # orphelins, lien dossier_cree retiré). Ces résidus ne sont pas de vrais
        # dossiers reçus.
        soumis = qs.exclude(statut=Dossier.Statut.BROUILLON)
        residus = soumis.filter(deposant__isnull=True).annotate(
            _lie=Exists(ReclamationEligibilite.objects.filter(dossier_cree=OuterRef('pk')))
        ).filter(_lie=False).count()
        return Response({
            'total': total,
            'par_statut': par_statut,
            'residus': residus,
            'recus_reels': soumis.count() - residus,
            'par_origine': {
                'reclamation': nb_reclam_val,
                'en_ligne': nb_retenu - nb_reclam_val,
            },
        })

    @action(detail=False, methods=['get'], url_path='stats-correspondance')
    def stats_correspondance(self, request):
        """Histogramme : brouillons + dossiers DÉPOSÉS ventilés par correspondance
        avec la liste d'éligibilité (rattaché / à rattacher / partielle / aucune).

        Réservé au back-office. Filtrable par `?appel=`. Les buckets reprennent
        exactement la logique des filtres de la liste (cf. `get_queryset`)."""
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")

        qs = Dossier.objects.all()
        appel = request.query_params.get('appel')
        if appel:
            qs = qs.filter(appel_id=appel)

        brouillon = qs.filter(statut=Dossier.Statut.BROUILLON).count()
        deposes = self._annoter_correspondance(qs.filter(statut=Dossier.Statut.DEPOSE))
        rattache = deposes.filter(ligne_eligibilite__isnull=False).count()
        reste = deposes.filter(ligne_eligibilite__isnull=True)
        a_rattacher = reste.filter(corresp_nom_complet=True).count()
        # Partielle = au moins un champ de NOM (le code seul ne compte pas).
        partielle = reste.filter(corresp_nom_complet=False).filter(
            Q(corresp_f_nom=True) | Q(corresp_f_postnom=True)
            | Q(corresp_f_prenom=True)
        ).count()
        aucune = reste.filter(
            corresp_nom_complet=False, corresp_f_nom=False,
            corresp_f_postnom=False, corresp_f_prenom=False,
        ).count()
        return Response({
            'brouillon': brouillon,
            'depose': {
                'rattache': rattache, 'a_rattacher': a_rattacher,
                'partielle': partielle, 'aucune': aucune,
                'total': rattache + a_rattacher + partielle + aucune,
            },
        })

    @action(detail=False, methods=['get'], url_path='doublon-suivant')
    def doublon_suivant(self, request):
        """Prochain dossier DÉPOSÉ ayant un doublon, pour balayer les doublons.

        `?apres=<id>` : renvoie le suivant (id croissant) ; boucle au début sinon.
        `?appel=<id>` : restreint à un appel. Renvoie {id, total}.
        """
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        base = (
            Dossier.objects
            .filter(statut=Dossier.Statut.DEPOSE)
            .annotate(a_doublon=self._doublon_exists())
            .filter(a_doublon=True)
            .order_by('id')
        )
        appel = request.query_params.get('appel')
        if appel:
            base = base.filter(appel_id=appel)
        total = base.count()
        apres = request.query_params.get('apres')
        suivant = base.filter(id__gt=apres).first() if apres else None
        if suivant is None:
            suivant = base.first()   # boucle au début
        return Response({'id': suivant.id if suivant else None, 'total': total})

    @action(detail=True, methods=['get', 'post'])
    def pieces(self, request, pk=None):
        """GET : liste les pièces. POST : ajoute une pièce (brouillon, déposant)."""
        dossier = self.get_object()
        if request.method == 'GET':
            return Response(
                PieceJointeSerializer(dossier.pieces.all(), many=True).data
            )

        self._verifier_modifiable(dossier)
        serializer = PieceJointeUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        type_piece = serializer.validated_data['type_piece']

        # Pièce « unique » : on refuse un second fichier du même type (sauf si
        # l'AAC autorise plusieurs fichiers pour ce type, ex. les diplômes).
        exigence = dossier.appel.pieces_exigees.filter(type_piece=type_piece).first()
        autorise_plusieurs = exigence.multiple if exigence else False
        if not autorise_plusieurs and dossier.pieces.filter(type_piece=type_piece).exists():
            raise ValidationError(
                "Cette pièce n'accepte qu'un seul fichier. Retirez l'actuel pour le remplacer."
            )

        fichier = serializer.validated_data['fichier']
        piece = serializer.save(
            dossier=dossier,
            nom_original=fichier.name[:255],
            taille=fichier.size,
        )
        return Response(
            PieceJointeSerializer(piece).data, status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=['delete'],
            url_path=r'pieces/(?P<piece_id>[^/.]+)')
    def supprimer_piece(self, request, pk=None, piece_id=None):
        """Retire une pièce d'un dossier encore en brouillon."""
        dossier = self.get_object()
        self._verifier_modifiable(dossier)
        piece = get_object_or_404(PieceJointe, pk=piece_id, dossier=dossier)
        piece.fichier.delete(save=False)  # supprime le fichier physique
        piece.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['get'],
            url_path=r'pieces/(?P<piece_id>[^/.]+)/telecharger')
    def telecharger_piece(self, request, pk=None, piece_id=None):
        """Téléchargement protégé d'une pièce (jamais d'URL publique).

        `?inline=1` sert le fichier en ligne (aperçu PDF/image dans l'app) au
        lieu de forcer le téléchargement.
        """
        dossier = self.get_object()  # déjà scopé par rôle/propriété
        piece = get_object_or_404(PieceJointe, pk=piece_id, dossier=dossier)
        try:
            fichier = piece.fichier.open('rb')
        except FileNotFoundError as exc:
            raise Http404("Fichier introuvable.") from exc
        inline = request.query_params.get('inline') in ('1', 'true', 'oui')
        reponse = FileResponse(fichier, as_attachment=not inline,
                               filename=piece.nom_original or 'piece')
        if inline:
            # Autorise l'affichage dans une iframe même origine (aperçu in-app).
            reponse['X-Frame-Options'] = 'SAMEORIGIN'
        return reponse

    @action(detail=True, methods=['post'])
    def soumettre(self, request, pk=None):
        """BROUILLON → DÉPOSÉ : verrouille le dossier si toutes les pièces

        obligatoires sont présentes, puis envoie l'accusé de réception.
        """
        dossier = self.get_object()
        if dossier.deposant_id != request.user.id and not request.user.is_superuser:
            raise PermissionDenied("Ce dossier ne vous appartient pas.")
        # Candidatures clôturées entre la création du brouillon et sa soumission :
        # on refuse le dépôt (cohérent avec le masquage du bouton « Postuler »).
        if not dossier.appel.est_ouvert:
            raise ValidationError(
                {'detail': "Les candidatures pour cet appel sont clôturées."}
            )
        manquantes = dossier.pieces_obligatoires_manquantes()
        if manquantes:
            raise ValidationError({
                'pieces_manquantes': [tp.libelle for tp in manquantes],
                'detail': "Des pièces obligatoires sont manquantes.",
            })
        self._rattacher_par_nom(dossier)
        try:
            # Verrou : un double-clic sur « Soumettre » ne dépose qu'une fois
            # (le second relit le statut DÉPOSÉ et échoue sans second accusé).
            with transaction.atomic():
                dossier = Dossier.objects.select_for_update().get(pk=dossier.pk)
                dossier.changer_statut(
                    Dossier.Statut.DEPOSE, par=request.user,
                    motif='Soumission par le candidat',
                )
        except DjangoValidationError as exc:
            raise ValidationError({'detail': exc.messages})

        self._envoyer_accuse(dossier)
        return Response(self.get_serializer(dossier).data)

    @staticmethod
    def _rattacher_par_nom(dossier):
        """Rattache la ligne d'éligibilité qui désigne cette même personne.

        Critère = nom complet identique (nom ET postnom ET prénom ; cf.
        `Dossier.ligne_eligibilite_correspondante`), JAMAIS le code seul : des
        candidats saisissent le code d'autrui (triche). Best-effort : seulement
        si le dossier n'est pas déjà rattaché et si la correspondance est unique.
        """
        if dossier.ligne_eligibilite_id:
            return
        ligne = dossier.ligne_eligibilite_correspondante()
        if ligne:
            dossier.ligne_eligibilite = ligne
            dossier.save(update_fields=['ligne_eligibilite'])

    @staticmethod
    def _code_dossier(dossier):
        """Code du dossier pour l'affichage (emails, sujets) ; repli sur #id."""
        return dossier.code or f'#{dossier.pk}'

    def _envoyer_accuse(self, dossier):
        """Accusé de réception au candidat (best-effort : n'annule pas le dépôt)."""
        code = self._code_dossier(dossier)
        try:
            envoyer_email(
                destinataire=dossier.email,
                sujet=f'Accusé de réception — dossier {code}',
                template='accuse_reception.html',
                contexte={
                    'nom_candidat': f'{dossier.nom} {dossier.postnom} {dossier.prenom}'.strip(),
                    'code_dossier': code,
                    'appel': dossier.appel.titre,
                },
            )
        except Exception:  # noqa: BLE001 — l'échec email ne doit pas casser le dépôt
            pass

    def _transition(self, request, vers, exige_role, motif_obligatoire=False,
                    lier_eligibilite=False, verif_validateur=False,
                    verif_affecte=False, email=None):
        """Factorise les actions de transition.

        :param lier_eligibilite: si True, rattache le dossier à la ligne de la
            liste d'éligibilité passée en `eligibilite_id` (traçabilité admin).
        :param verif_validateur: si True, exige que l'utilisateur soit désigné
            ET autorisé à valider ce dossier précis.
        :param verif_affecte: si True, exige que le dossier soit affecté à
            l'utilisateur (sauf admin) — verrou de répartition.
        :param email: tuple (sujet, template) à notifier au candidat après la
            transition (best-effort, n'annule pas la transition en cas d'échec).
        """
        if not exige_role(request.user):
            raise PermissionDenied(
                "Votre rôle ne vous autorise pas cette action."
            )

        corps = ChangementStatutSerializer(data=request.data)
        corps.is_valid(raise_exception=True)
        motif = corps.validated_data['motif']
        if motif_obligatoire and not motif.strip():
            raise ValidationError({'motif': "Un motif est obligatoire."})

        dossier = self.get_object()

        if verif_validateur:
            self._verifier_validateur(dossier)
        if verif_affecte:
            self._verifier_affecte(dossier)

        # Verrou ligne par ligne : deux décisions simultanées sur le MÊME
        # dossier (double-clic, deux agents) se sérialisent ; la seconde relit
        # le statut à jour et échoue proprement si la transition est devenue
        # interdite. Ne touche jamais qu'un seul dossier (pk de l'URL).
        try:
            with transaction.atomic():
                dossier = Dossier.objects.select_for_update().get(pk=dossier.pk)
                if lier_eligibilite:
                    eid = request.data.get('eligibilite_id')
                    if eid:
                        ligne = get_object_or_404(ListeEligibilite, pk=eid)
                        dossier.ligne_eligibilite = ligne
                        dossier.save(update_fields=['ligne_eligibilite'])
                dossier.changer_statut(vers, par=request.user, motif=motif)
        except DjangoValidationError as exc:
            raise ValidationError({'detail': exc.messages})

        if email:
            self._notifier(dossier, email, motif)

        return Response(self.get_serializer(dossier).data, status=status.HTTP_200_OK)

    def _notifier(self, dossier, email, motif=''):
        """Notifie le candidat d'un changement de statut (best-effort)."""
        sujet, template = email
        try:
            envoyer_email(
                destinataire=dossier.email,
                sujet=sujet,
                template=template,
                contexte={
                    'nom_candidat': f'{dossier.nom} {dossier.postnom} {dossier.prenom}'.strip(),
                    'code_dossier': self._code_dossier(dossier),
                    'appel': dossier.appel.titre,
                    'motif': motif,
                },
            )
        except Exception:  # noqa: BLE001 — l'échec email ne casse pas la transition
            pass

    # --- Actions ADMIN (dossier DÉPOSÉ) ---------------------------------

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """DÉPOSÉ → RETENU en une étape (validation directe back-office).

        Enchaîne DÉPOSÉ → EN_EXAMEN → RETENU via `changer_statut` (la machine à
        états et l'audit sont préservés : deux entrées d'historique). C'est le
        circuit en 1 clic demandé pour le traitement courant (l'étape « examen »
        n'est pas utilisée). Réservé à l'agent affecté ou à un admin.

        `eligibilite_id` optionnel : rattache la ligne d'éligibilité choisie ;
        à défaut, rattachement best-effort par nom complet. Aucun email pendant
        le traitement (le retenu est notifié à la publication de la liste)."""
        if not roles.peut_traiter(request.user):
            raise PermissionDenied("Réservé aux administrateurs et validateurs.")
        dossier = self.get_object()
        self._verifier_affecte(dossier)
        try:
            with transaction.atomic():
                dossier = Dossier.objects.select_for_update().get(pk=dossier.pk)
                # À valider = DÉPOSÉ ou EN_EXAMEN (l'examen n'étant plus une
                # étape distincte, on accepte les deux comme point de départ).
                if dossier.statut not in (Dossier.Statut.DEPOSE, Dossier.Statut.EN_EXAMEN):
                    raise ValidationError("Ce dossier n'est plus en attente de validation.")
                # Correction d'identité éventuelle : c'est ce NOM/POSTNOM/PRÉNOM
                # qui sera publié dans la liste des retenus. Permet de retenir
                # une personne hors liste (sans correspondance) avec un nom propre.
                identite_modifiee = False
                for champ in ('nom', 'postnom', 'prenom'):
                    if champ in request.data:
                        val = (request.data.get(champ) or '').strip()
                        if champ in ('nom', 'prenom') and not val:
                            raise ValidationError({champ: "Ce champ ne peut pas être vide."})
                        if getattr(dossier, champ) != val:
                            setattr(dossier, champ, val)
                            identite_modifiee = True
                # Correction éventuelle du poste visé (mal renseigné).
                if 'poste_id' in request.data:
                    pid = request.data.get('poste_id')
                    nouveau_poste = get_object_or_404(Poste, pk=pid) if pid else None
                    if dossier.poste_id != (nouveau_poste.id if nouveau_poste else None):
                        dossier.poste = nouveau_poste
                        identite_modifiee = True
                eid = request.data.get('eligibilite_id')
                if eid:
                    dossier.ligne_eligibilite = get_object_or_404(ListeEligibilite, pk=eid)
                    identite_modifiee = True   # déclenche un save complet ci-dessous
                if identite_modifiee:
                    dossier.save()   # save() recalcule texte_recherche
                elif not dossier.ligne_eligibilite_id:
                    # Aucune correction ni correspondance choisie : rattachement
                    # best-effort par nom exact (sinon on retient sans lien).
                    self._rattacher_par_nom(dossier)
                # Passe par EN_EXAMEN seulement si on part de DÉPOSÉ (audit).
                if dossier.statut == Dossier.Statut.DEPOSE:
                    dossier.changer_statut(Dossier.Statut.EN_EXAMEN, par=request.user,
                                           motif='Validation directe (back-office)')
                dossier.changer_statut(Dossier.Statut.RETENU, par=request.user,
                                       motif='Validation directe (back-office)')
        except DjangoValidationError as exc:
            raise ValidationError({'detail': exc.messages})
        return Response(self.get_serializer(dossier).data)

    @action(detail=True, methods=['post'])
    def approuver(self, request, pk=None):
        """DÉPOSÉ → EN_EXAMEN (admin ou validateur). `eligibilite_id` optionnel.

        Conservé pour le circuit en 2 étapes (examen évaluateur). Le traitement
        courant passe par `valider` (1 clic). Aucun email pendant le traitement.
        """
        return self._transition(
            request, Dossier.Statut.EN_EXAMEN, roles.peut_traiter,
            lier_eligibilite=True, verif_affecte=True,
        )

    @action(detail=True, methods=['post'])
    def rejeter(self, request, pk=None):
        """DÉPOSÉ → REJETÉ (admin ou validateur, motif obligatoire). Aucun email."""
        return self._transition(
            request, Dossier.Statut.REJETE, roles.peut_traiter,
            motif_obligatoire=True, verif_affecte=True,
        )

    @action(detail=True, methods=['post'], url_path='rejeter-doublon')
    def rejeter_doublon(self, request, pk=None):
        """Rejette un dossier en double (DÉPOSÉ → REJETÉ), sans email.

        Raccourci pour traiter les doublons : motif « Dossier en double »
        pré-rempli, pas de notification au candidat.
        """
        if not roles.peut_traiter(request.user):
            raise PermissionDenied("Réservé aux administrateurs et validateurs.")
        dossier = self.get_object()
        self._verifier_affecte(dossier)
        try:
            with transaction.atomic():
                dossier = Dossier.objects.select_for_update().get(pk=dossier.pk)
                dossier.changer_statut(
                    Dossier.Statut.REJETE, par=request.user, motif='Dossier en double',
                )
        except DjangoValidationError as exc:
            raise ValidationError({'detail': exc.messages})
        return Response(self.get_serializer(dossier).data)

    # --- Affectation des évaluateurs (admin) ----------------------------

    @action(detail=True, methods=['get', 'post'])
    def affectations(self, request, pk=None):
        """GET : liste les évaluateurs désignés. POST (admin) : en désigne un.

        Corps POST : { evaluateur_id, peut_valider (bool, défaut False) }.
        Idempotent : ré-affecter met à jour le droit de validation.
        """
        dossier = self.get_object()
        if request.method == 'GET':
            # Consultation des désignations : back-office ou évaluateur désigné.
            # Jamais le candidat (ne doit pas connaître ses évaluateurs).
            if not roles.acces_backoffice(request.user):
                self._verifier_designe(dossier)
            return Response(
                AffectationSerializer(dossier.affectations.all(), many=True).data
            )

        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Seul un administrateur ou un superviseur peut désigner un évaluateur.")
        evaluateur = get_object_or_404(User, pk=request.data.get('evaluateur_id'))
        if not roles.est_evaluateur(evaluateur):
            raise ValidationError(
                {'evaluateur_id': "Cet utilisateur n'est pas un évaluateur."}
            )
        affectation, _ = AffectationEvaluateur.objects.update_or_create(
            dossier=dossier, evaluateur=evaluateur,
            defaults={'peut_valider': bool(request.data.get('peut_valider', False))},
        )
        return Response(
            AffectationSerializer(affectation).data, status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=['delete'],
            url_path=r'affectations/(?P<evaluateur_id>[^/.]+)')
    def retirer_affectation(self, request, pk=None, evaluateur_id=None):
        """Retire la désignation d'un évaluateur (admin/superviseur)."""
        dossier = self.get_object()
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Seul un administrateur ou un superviseur peut retirer une désignation.")
        affectation = get_object_or_404(
            AffectationEvaluateur, dossier=dossier, evaluateur_id=evaluateur_id,
        )
        affectation.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # --- Évaluation (évaluateur désigné) --------------------------------

    def _verifier_designe(self, dossier):
        """L'utilisateur est-il désigné sur ce dossier (ou admin/superviseur) ?"""
        user = self.request.user
        if roles.peut_superviser(user):
            return
        if not dossier.affectations.filter(evaluateur=user).exists():
            raise PermissionDenied("Vous n'êtes pas désigné sur ce dossier.")

    def _verifier_affecte(self, dossier):
        """Verrou d'affectation pour traiter un dossier (approuver/rejeter).

        Admin et superviseur : toujours (ils tranchent et réaffectent, comme
        pour les réclamations). Validateur : seulement si le dossier LUI est
        affecté (évite de traiter le lot d'un collègue)."""
        user = self.request.user
        if roles.peut_superviser(user):
            return
        if dossier.affecte_a_id != user.id:
            raise PermissionDenied(
                "Ce dossier est affecté à un autre agent (ou pas encore affecté)."
            )

    def _verifier_validateur(self, dossier):
        """Qui peut trancher retenir/non-retenir : l'admin ou le superviseur
        (toujours), un évaluateur désigné ET autorisé (peut_valider), ou le
        validateur à qui le dossier est affecté."""
        user = self.request.user
        if roles.peut_superviser(user):
            return
        # Évaluateur désigné et autorisé : indépendant de l'affectation.
        if dossier.affectations.filter(evaluateur=user, peut_valider=True).exists():
            return
        # Validateur : seulement son propre lot.
        if roles.est_validateur(user) and dossier.affecte_a_id == user.id:
            return
        raise PermissionDenied(
            "Vous devez être l'agent affecté à ce dossier, ou un évaluateur autorisé."
        )

    @action(detail=True, methods=['get', 'post'])
    def evaluations(self, request, pk=None):
        """GET : liste les avis. POST : enregistre/mention l'avis de l'évaluateur.

        Corps POST : { avis, recommandation }. Un évaluateur a un seul avis par
        dossier (mis à jour s'il rejoue).
        """
        dossier = self.get_object()
        if request.method == 'GET':
            # Consultation des avis : tout le back-office, ou un désigné.
            if not roles.acces_backoffice(request.user):
                self._verifier_designe(dossier)
            return Response(
                EvaluationSerializer(dossier.evaluations.all(), many=True).data
            )

        self._verifier_designe(dossier)
        evaluation = Evaluation.objects.filter(
            dossier=dossier, evaluateur=request.user,
        ).first()
        serializer = EvaluationSerializer(evaluation, data=request.data, partial=bool(evaluation))
        serializer.is_valid(raise_exception=True)
        serializer.save(dossier=dossier, evaluateur=request.user)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # --- Décision (évaluateur désigné ET autorisé, dossier EN_EXAMEN) ----

    # Décision finale : l'admin ou un validateur tranche, ou un évaluateur
    # désigné autorisé (vérifié ensuite dossier par dossier).
    _peut_decider = staticmethod(
        lambda u: roles.peut_traiter(u) or roles.est_evaluateur(u)
    )

    @action(detail=True, methods=['post'])
    def retenir(self, request, pk=None):
        """EN_EXAMEN → RETENU (admin, ou évaluateur désigné et autorisé).

        Aucun email ici : le retenu est notifié à la publication de la liste.
        """
        return self._transition(
            request, Dossier.Statut.RETENU, self._peut_decider,
            verif_validateur=True,
        )

    @action(detail=True, methods=['post'], url_path='non-retenir')
    def non_retenir(self, request, pk=None):
        """EN_EXAMEN → NON_RETENU (admin ou évaluateur autorisé, motif requis).
        Aucun email (traitement)."""
        return self._transition(
            request, Dossier.Statut.NON_RETENU, self._peut_decider,
            motif_obligatoire=True, verif_validateur=True,
        )

    @action(detail=True, methods=['post'])
    def rouvrir(self, request, pk=None):
        """Rouvre un dossier DÉCIDÉ (retenu / non-retenu / rejeté) → DÉPOSÉ, pour
        correction (admin / superviseur).

        Le dossier repasse dans la file « à valider » : l'agent peut alors
        re-trancher (Valider / Rejeter) avec le bon statut et le bon motif, et
        corriger l'identité si besoin. La réouverture est tracée dans
        l'historique. Si le candidat le consultait, sa décision redevient
        masquée (le dossier n'est plus décidé)."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        try:
            with transaction.atomic():
                dossier = Dossier.objects.select_for_update().get(pk=self.get_object().pk)
                if dossier.statut not in Dossier.STATUTS_TERMINAUX:
                    raise ValidationError("Ce dossier n'est pas décidé (déjà à traiter).")
                ancien = dossier.statut
                dossier.statut = Dossier.Statut.DEPOSE
                dossier.save(update_fields=['statut', 'modifie_le'])
                HistoriqueStatut.objects.create(
                    dossier=dossier, ancien_statut=ancien,
                    nouveau_statut=Dossier.Statut.DEPOSE, par=request.user,
                    motif='Dossier rouvert pour correction',
                )
        except DjangoValidationError as exc:
            raise ValidationError({'detail': exc.messages})
        return Response(self.get_serializer(dossier).data)

    @action(detail=True, methods=['get'])
    def historique(self, request, pk=None):
        """Journal d'audit des changements de statut du dossier.

        - **Back-office** : historique COMPLET (toutes les étapes + auteurs).
        - **Candidat** : vue épurée — uniquement le PREMIER statut
          (Brouillon → Déposé) et la DÉCISION FINALE (retenu/non-retenu/rejeté)
          avec son motif, et seulement une fois les résultats publiés. Jamais
          les étapes intermédiaires, jamais les noms des utilisateurs."""
        dossier = self.get_object()
        qs = dossier.historique.all()

        if roles.acces_backoffice(request.user):
            return Response(HistoriqueStatutSerializer(qs, many=True).data)

        # Vue candidat : on ne retient que le 1er (dépôt) et la décision finale.
        entrees = []
        decision_publiee = (
            dossier.statut in Dossier.STATUTS_TERMINAUX
            and dossier.appel.liste_retenus_publiee
        )
        if decision_publiee:
            finale = (qs.filter(nouveau_statut__in=Dossier.STATUTS_TERMINAUX)
                      .order_by('horodatage').last())
            if finale:
                entrees.append(finale)   # décision (la plus récente d'abord)
        premier = (qs.filter(nouveau_statut=Dossier.Statut.DEPOSE)
                   .order_by('horodatage').first())
        if premier and premier not in entrees:
            entrees.append(premier)

        data = HistoriqueStatutSerializer(entrees, many=True).data
        for d in data:
            d['par'] = None   # jamais publier les noms des utilisateurs
        return Response(data)

    # --- Répartition de la charge entre agents --------------------------

    @action(detail=False, methods=['post'])
    def repartir(self, request):
        """Répartit équitablement les dossiers FILTRÉS entre des agents (admin).

        Corps : {agents: [id…], statut?, appel?, correspondance?, doublons?, q?,
        seulement_non_affectes?: bool}. Reprend EXACTEMENT les filtres de la
        liste (mêmes buckets d'éligibilité via `_filtrer_liste`) : on peut donc
        répartir une seule catégorie (ex. « aucune correspondance ») entre les
        agents. Par défaut statut = DÉPOSÉ (la file à valider) et seuls les
        dossiers non encore affectés sont distribués ; `seulement_non_affectes`
        à False = RÉÉQUILIBRAGE (réaffecte aussi les déjà affectés). L'agent
        affecté mène le dossier de l'approbation à la décision.
        Opération additive : ne change que `affecte_a`."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        agent_ids = request.data.get('agents') or []
        if not isinstance(agent_ids, list) or not agent_ids:
            raise ValidationError({'agents': "Sélectionnez au moins un agent."})

        # Catégorie déjà décidée (retenu/non-retenu/rejeté) → révision réservée
        # aux superviseurs ; file « à valider » (déposé/examen) → agents de
        # traitement (validateur/superviseur).
        statut_param = request.data.get('statut') or Dossier.Statut.DEPOSE
        statuts = [s for s in str(statut_param).split(',') if s]
        en_cours = {Dossier.Statut.DEPOSE, Dossier.Statut.EN_EXAMEN}
        cible_decidee = any(s not in en_cours for s in statuts)
        eligible = roles.peut_superviser if cible_decidee else roles.peut_traiter

        trouves = {u.id: u for u in User.objects.filter(id__in=agent_ids, is_active=True)}
        agents = [trouves[i] for i in agent_ids if i in trouves and eligible(trouves[i])]
        if not agents:
            raise ValidationError({'agents': (
                "Pour une catégorie déjà décidée (retenus / non-retenus / rejetés), "
                "seuls des superviseurs peuvent être affectés." if cible_decidee else
                "Aucun agent valide (validateur ou superviseur actif).")})

        qs = self._annoter_correspondance(Dossier.objects.all()).annotate(
            a_doublon=self._doublon_exists(),
        )
        filtres = {
            'statut': statut_param,
            'appel': request.data.get('appel'),
            'correspondance': request.data.get('correspondance'),
            'doublons': request.data.get('doublons'),
            'q': request.data.get('q'),
        }
        qs = self._filtrer_liste(qs, filtres)
        if request.data.get('seulement_non_affectes', True):
            qs = qs.filter(affecte_a__isnull=True)
        ids = list(qs.order_by('cree_le').values_list('id', flat=True))

        par_agent = {u.id: [] for u in agents}
        for i, did in enumerate(ids):
            par_agent[agents[i % len(agents)].id].append(did)

        with transaction.atomic():
            for u in agents:
                lot = par_agent[u.id]
                if lot:
                    Dossier.objects.filter(id__in=lot).update(affecte_a=u)

        return Response({
            'total_reparti': len(ids),
            'par_agent': [
                {'agent_id': u.id, 'agent': (u.get_full_name() or u.email),
                 'attribues': len(par_agent[u.id])}
                for u in agents
            ],
        })

    @action(detail=False, methods=['get'])
    def repartition(self, request):
        """Charge par agent : total affecté + non encore tranché (back-office).

        « a_traiter » = dossiers affectés encore DÉPOSÉ ou EN_EXAMEN (en cours)."""
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        en_cours = [Dossier.Statut.DEPOSE, Dossier.Statut.EN_EXAMEN]
        lignes = (
            Dossier.objects
            .filter(affecte_a__isnull=False)
            .values('affecte_a_id', 'affecte_a__first_name',
                    'affecte_a__last_name', 'affecte_a__email')
            .annotate(
                total=Count('id'),
                a_traiter=Count('id', filter=Q(statut__in=en_cours)),
            )
            .order_by('affecte_a__first_name', 'affecte_a__last_name')
        )
        resultat = []
        for l in lignes:
            nom = f"{l['affecte_a__first_name']} {l['affecte_a__last_name']}".strip()
            resultat.append({
                'agent_id': l['affecte_a_id'],
                'agent': nom or l['affecte_a__email'],
                'total': l['total'],
                'a_traiter': l['a_traiter'],
                'traites': l['total'] - l['a_traiter'],
            })
        return Response({'par_agent': resultat})


class ReclamationThrottle(AnonRateThrottle):
    """Limite le formulaire public de réclamation (anti-spam)."""

    scope = 'reclamation'


class CritereValidationViewSet(viewsets.ReadOnlyModelViewSet):
    """Grille de critères de validation (lecture, back-office).

    La configuration (ajout/édition/suppression) se fait dans la console
    Django. Le front ne fait que lister les critères ACTIFS, filtrables par
    portée : `GET /api/criteres/?portee=reclamation`.
    """

    serializer_class = CritereValidationSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None   # petite liste de config : tableau simple

    def get_queryset(self):
        if not roles.acces_backoffice(self.request.user):
            return CritereValidation.objects.none()
        qs = CritereValidation.objects.filter(actif=True)
        portee = self.request.query_params.get('portee')
        if portee:
            # Inclut les critères « les deux ».
            qs = qs.filter(portee__in=[portee, CritereValidation.Portee.LES_DEUX])
        return qs


class ReclamationViewSet(viewsets.ModelViewSet):
    """Réclamations d'éligibilité (personne absente de la liste).

    - **création** : publique (sans compte), throttlée + honeypot ;
    - **liste / détail / actions / téléchargement** : réservés aux administrateurs.

    À la validation, un `Dossier` est créé et conduit jusqu'à RETENU via
    `Dossier.changer_statut()` (l'audit et l'invariant de statut sont préservés).
    """

    # 25 par page, taille ajustable (page_size) — comme la file de validation.
    pagination_class = PaginationStandard
    # Multipart pour la création (fichiers) ; JSON pour valider/rejeter.
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_serializer_class(self):
        if self.action == 'create':
            return ReclamationCreationSerializer
        return ReclamationAdminSerializer

    def get_permissions(self):
        return [AllowAny()] if self.action == 'create' else [IsAuthenticated()]

    def get_throttles(self):
        return [ReclamationThrottle()] if self.action == 'create' else []

    def get_queryset(self):
        # Hors création, l'accès est réservé au back-office (lecture ; les
        # actions valider/rejeter sont contrôlées séparément).
        if not roles.acces_backoffice(self.request.user):
            return ReclamationEligibilite.objects.none()
        qs = ReclamationEligibilite.objects.select_related(
            'appel', 'poste', 'traite_par', 'affecte_a', 'dossier_cree',
        ).prefetch_related('documents', 'controles').annotate(
            # Doublon probable : une AUTRE réclamation du même appel, même nom
            # complet (normalisé), pas encore rejetée. Indicatif.
            a_doublon=Exists(
                ReclamationEligibilite.objects
                .filter(appel_id=OuterRef('appel_id'))
                .exclude(statut=ReclamationEligibilite.Statut.REJETEE)
                .exclude(texte_recherche='')
                .filter(texte_recherche=OuterRef('texte_recherche'))
                .exclude(pk=OuterRef('pk'))
            ),
            # Croisement : la personne a-t-elle DÉJÀ un dossier DÉPOSÉ (même nom
            # complet normalisé) ? Sa réclamation est alors redondante (elle est
            # déjà candidate). Par nom, jamais l'email (cf. règle anti-triche).
            a_dossier_depose=Exists(
                Dossier.objects
                .filter(statut=Dossier.Statut.DEPOSE)
                .exclude(texte_recherche='')
                .filter(texte_recherche=OuterRef('texte_recherche'))
            ),
        )
        statut = self.request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        appel = self.request.query_params.get('appel')
        if appel:
            qs = qs.filter(appel_id=appel)
        if self.request.query_params.get('dossier_depose') in ('1', 'true', 'oui'):
            qs = qs.filter(a_dossier_depose=True)
        # Réclamations sans poste renseigné (à corriger).
        if self.request.query_params.get('sans_poste') in ('1', 'true', 'oui'):
            qs = qs.filter(poste__isnull=True)
        # Filtre d'affectation : « moi » = mon lot, « aucune » = non affectées,
        # un id = le lot d'un agent (pour l'admin qui supervise la répartition).
        affecte = self.request.query_params.get('affecte')
        if affecte == 'moi':
            qs = qs.filter(affecte_a=self.request.user)
        elif affecte == 'aucune':
            qs = qs.filter(affecte_a__isnull=True)
        elif affecte and affecte.isdigit():
            qs = qs.filter(affecte_a_id=int(affecte))
        q = self.request.query_params.get('q')
        if q:
            for token in tokens_recherche(q):
                qs = qs.filter(texte_recherche__contains=token)

        # Tri demandé par le tableau (allowlist) ; sinon plus récentes d'abord.
        ordering = self.request.query_params.get('ordering', '')
        if ordering.lstrip('-') in self.TRI_AUTORISE:
            qs = qs.order_by(ordering)
        else:
            qs = qs.order_by('-cree_le')
        return qs

    # Champs autorisés au tri (allowlist : évite l'injection de champ arbitraire).
    TRI_AUTORISE = {
        'id', 'nom', 'email', 'statut', 'cree_le',
        'appel__titre', 'affecte_a__first_name',
    }

    # Documents attendus : accusé, CV, pièce d'identité (un chacun) + diplôme(s).
    DOCS_SIMPLES = [
        ('accuse', DocumentReclamation.Type.ACCUSE, "l'accusé de réception"),
        ('cv', DocumentReclamation.Type.CV, 'le CV'),
        ('identite', DocumentReclamation.Type.IDENTITE, "la pièce d'identité"),
    ]

    def _valider_fichier(self, fichier, libelle):
        ext = fichier.name.rsplit('.', 1)[-1].lower() if '.' in fichier.name else ''
        if ext not in EXTENSIONS_AUTORISEES:
            raise ValidationError(
                {'detail': f"Format non autorisé pour {libelle} (acceptés : "
                           f"{', '.join(EXTENSIONS_AUTORISEES)})."}
            )
        if fichier.size > TAILLE_MAX_PIECE:
            limite = TAILLE_MAX_PIECE // (1024 * 1024)
            raise ValidationError({'detail': f"Fichier trop volumineux pour {libelle} (max {limite} Mo)."})

    def create(self, request, *args, **kwargs):
        """Dépôt public : valide texte + justificatifs, réponse neutre.

        Justificatifs requis : accusé de réception, CV, pièce d'identité (un
        chacun) et au moins un diplôme (plusieurs possibles).
        """
        serializer = self.get_serializer(data=request.data)
        # `validate_appel` (serializer) refuse déjà un appel non publié : la
        # réclamation est donc fermée en même temps que les candidatures.
        serializer.is_valid(raise_exception=True)

        # Collecte et validation des fichiers.
        simples = {}
        for champ, type_doc, libelle in self.DOCS_SIMPLES:
            fichier = request.FILES.get(champ)
            if not fichier:
                raise ValidationError({'detail': f"Veuillez joindre {libelle}."})
            self._valider_fichier(fichier, libelle)
            simples[type_doc] = fichier

        diplomes = request.FILES.getlist('diplomes')
        if not diplomes:
            raise ValidationError({'detail': "Veuillez joindre au moins un diplôme (ou équivalent)."})
        for d in diplomes:
            self._valider_fichier(d, 'un diplôme')

        with transaction.atomic():
            reclamation = serializer.save()
            for type_doc, fichier in simples.items():
                DocumentReclamation.objects.create(
                    reclamation=reclamation, type=type_doc, fichier=fichier,
                    nom_original=fichier.name[:255], taille=fichier.size,
                )
            for fichier in diplomes:
                DocumentReclamation.objects.create(
                    reclamation=reclamation, type=DocumentReclamation.Type.DIPLOME,
                    fichier=fichier, nom_original=fichier.name[:255], taille=fichier.size,
                )

        self._accuse_reception(reclamation)
        return Response(
            {'detail': "Votre réclamation a bien été enregistrée. "
                       "Vous recevrez une réponse par email."},
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Comptes de réclamations par statut (back-office), pour les KPI."""
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        par_statut = {
            row['statut']: row['n']
            for row in ReclamationEligibilite.objects.values('statut').annotate(n=Count('id'))
        }
        return Response({'total': sum(par_statut.values()), 'par_statut': par_statut})

    @action(detail=True, methods=['get'])
    def doublons(self, request, pk=None):
        """Autres réclamations du même appel et même nom complet (non rejetées)."""
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        reclamation = self.get_object()
        if not reclamation.texte_recherche:
            return Response([])
        autres = (
            ReclamationEligibilite.objects
            .filter(appel_id=reclamation.appel_id, texte_recherche=reclamation.texte_recherche)
            .exclude(statut=ReclamationEligibilite.Statut.REJETEE)
            .exclude(pk=reclamation.pk)
            .prefetch_related('documents')
            .order_by('cree_le')
        )
        return Response([
            {
                'id': r.id, 'nom': r.nom, 'postnom': r.postnom, 'prenom': r.prenom,
                'email': r.email, 'statut': r.statut,
                'statut_libelle': r.get_statut_display(), 'cree_le': r.cree_le,
                # Justificatifs du doublon : permet de vérifier avant de rejeter.
                'documents': [
                    {
                        'id': d.id, 'type': d.type,
                        'type_libelle': d.get_type_display(),
                        'nom_original': d.nom_original, 'taille': d.taille,
                    }
                    for d in r.documents.all()
                ],
            }
            for r in autres
        ])

    @action(detail=True, methods=['get'], url_path='dossiers-deposes')
    def dossiers_deposes(self, request, pk=None):
        """Dossiers DÉPOSÉS portant le même nom complet (personne déjà candidate).

        Sert à décider : si la personne a déjà un dossier déposé, sa réclamation
        est redondante et peut être rejetée. Rapprochement par nom (jamais email).
        """
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        reclamation = self.get_object()
        if not reclamation.texte_recherche:
            return Response([])
        dossiers = (
            Dossier.objects
            .filter(statut=Dossier.Statut.DEPOSE, texte_recherche=reclamation.texte_recherche)
            .select_related('appel', 'poste', 'affecte_a')
            .prefetch_related('pieces__type_piece')
            .order_by('cree_le')
        )
        return Response([
            {
                'id': d.id, 'code': d.code, 'nom': d.nom, 'postnom': d.postnom,
                'prenom': d.prenom, 'statut': d.statut,
                'statut_libelle': d.get_statut_display(),
                'appel_titre': d.appel.titre,
                'poste_libelle': d.poste.libelle if d.poste else None,
                'affecte_a': d.affecte_a_id,
                'affecte_a_nom': (d.affecte_a.get_full_name() or d.affecte_a.email)
                if d.affecte_a else None,
                'cree_le': d.cree_le,
                # Pièces du dossier : permet de décider sans ouvrir le dossier.
                'pieces': [
                    {
                        'id': p.id,
                        'type_libelle': p.type_piece.libelle if p.type_piece else 'Pièce',
                        'nom_original': p.nom_original, 'taille': p.taille,
                    }
                    for p in d.pieces.all()
                ],
            }
            for d in dossiers
        ])

    @action(detail=False, methods=['get'], url_path='doublon-suivant')
    def doublon_suivant(self, request):
        """Prochaine réclamation EN ATTENTE ayant un doublon (pour les balayer).

        `?apres=<id>` renvoie la suivante ; boucle au début sinon. {id, total}.
        """
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        sous_req = (
            ReclamationEligibilite.objects
            .filter(appel_id=OuterRef('appel_id'))
            .exclude(statut=ReclamationEligibilite.Statut.REJETEE)
            .exclude(texte_recherche='')
            .filter(texte_recherche=OuterRef('texte_recherche'))
            .exclude(pk=OuterRef('pk'))
        )
        base = (
            ReclamationEligibilite.objects
            .filter(statut=ReclamationEligibilite.Statut.EN_ATTENTE)
            .annotate(a_doublon=Exists(sous_req))
            .filter(a_doublon=True)
            .order_by('id')
        )
        appel = request.query_params.get('appel')
        if appel:
            base = base.filter(appel_id=appel)
        total = base.count()
        apres = request.query_params.get('apres')
        suivant = base.filter(id__gt=apres).first() if apres else None
        if suivant is None:
            suivant = base.first()
        return Response({'id': suivant.id if suivant else None, 'total': total})

    @action(detail=True, methods=['get'],
            url_path=r'documents/(?P<doc_id>[^/.]+)')
    def telecharger_document(self, request, pk=None, doc_id=None):
        """Téléchargement protégé d'un justificatif (back-office ; jamais public)."""
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        reclamation = self.get_object()
        doc = get_object_or_404(DocumentReclamation, pk=doc_id, reclamation=reclamation)
        try:
            fichier = doc.fichier.open('rb')
        except FileNotFoundError as exc:
            raise Http404("Fichier introuvable.") from exc
        inline = request.query_params.get('inline') in ('1', 'true', 'oui')
        reponse = FileResponse(fichier, as_attachment=not inline,
                               filename=doc.nom_original or 'document')
        if inline:
            reponse['X-Frame-Options'] = 'SAMEORIGIN'
        return reponse

    def _verifier_peut_decider(self, user, reclamation):
        """Trancher = être admin, ou être le validateur à qui c'est affecté."""
        if not roles.peut_traiter(user):
            raise PermissionDenied("Réservé aux administrateurs et validateurs.")
        if not roles.peut_decider_affecte(user, reclamation.affecte_a_id):
            raise PermissionDenied("Cette réclamation est affectée à un autre agent.")

    @staticmethod
    def _criteres_reclamation():
        """Critères actifs applicables aux réclamations (réclamation + les deux)."""
        return list(CritereValidation.objects.filter(
            actif=True,
            portee__in=[CritereValidation.Portee.RECLAMATION,
                        CritereValidation.Portee.LES_DEUX],
        ))

    @staticmethod
    def _ids_coches(request):
        """Ids de critères cochés, envoyés par le front (tolérant aux types)."""
        coches = set()
        for x in (request.data.get('criteres') or []):
            try:
                coches.add(int(x))
            except (TypeError, ValueError):
                pass
        return coches

    @staticmethod
    def _enregistrer_grille(reclamation, actifs, coches, par):
        """Photographie de la grille (audit), avec copie du libellé. Utilisé
        aussi bien à la validation qu'au rejet : on garde la trace de ce que la
        personne remplit ou non, quelle que soit la décision."""
        for c in actifs:
            ControleCritere.objects.update_or_create(
                reclamation=reclamation, critere=c,
                defaults={'libelle_snapshot': c.libelle,
                          'rempli': c.id in coches, 'par': par},
            )

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """Valide la réclamation : crée un dossier et le conduit jusqu'à RETENU.

        Contrôle d'abord la grille de critères (portée réclamation) : tous les
        critères actifs doivent être cochés. Sinon, seul un administrateur peut
        forcer avec une dérogation justifiée (`derogation`). Les contrôles sont
        enregistrés (audit), avec copie du libellé."""
        reclamation = self.get_object()
        self._verifier_peut_decider(request.user, reclamation)

        # Poste : celui choisi par l'admin à la validation, sinon celui déclaré
        # par le réclamant dans le formulaire (absent sur les anciennes
        # réclamations, d'où le repli sur None).
        poste_id = request.data.get('poste_id')
        poste = get_object_or_404(Poste, pk=poste_id) if poste_id else reclamation.poste

        # Grille de validation (bloquant avec dérogation admin).
        actifs = self._criteres_reclamation()
        coches = self._ids_coches(request)
        manquants = [c for c in actifs if c.id not in coches]
        derogation = (request.data.get('derogation') or '').strip()
        if manquants and not (roles.est_admin(request.user) and derogation):
            raise ValidationError({
                'criteres_manquants': [c.libelle for c in manquants],
                'detail': "Critères non remplis : cochez-les, ou faites valider "
                          "par un administrateur avec une dérogation justifiée.",
            })

        with transaction.atomic():
            # Verrou + re-contrôle du statut SOUS verrou : deux validations
            # simultanées de la même réclamation ne peuvent pas créer deux
            # dossiers — la seconde échoue avec « déjà traitée ».
            reclamation = (
                ReclamationEligibilite.objects
                .select_for_update().get(pk=reclamation.pk)
            )
            if reclamation.statut != ReclamationEligibilite.Statut.EN_ATTENTE:
                raise ValidationError("Cette réclamation a déjà été traitée.")
            dossier = Dossier.objects.create(
                appel=reclamation.appel, poste=poste, deposant=None,
                nom=reclamation.nom, postnom=reclamation.postnom,
                prenom=reclamation.prenom, email=reclamation.email,
                statut=Dossier.Statut.BROUILLON,
            )
            # Dossier issu d'une réclamation : exempté des pièces obligatoires
            # (l'accusé de réception tient lieu de justificatif). On enchaîne les
            # transitions pour conserver l'audit et respecter la machine à états.
            dossier.changer_statut(Dossier.Statut.DEPOSE, par=request.user,
                                   motif='Validé via réclamation (accusé de réception ACGT)')
            dossier.changer_statut(Dossier.Statut.EN_EXAMEN, par=request.user,
                                   motif='Validé via réclamation')
            dossier.changer_statut(Dossier.Statut.RETENU, par=request.user,
                                   motif='Validé via réclamation')
            reclamation.statut = ReclamationEligibilite.Statut.VALIDEE
            reclamation.traite_par = request.user
            reclamation.traite_le = timezone.now()
            reclamation.dossier_cree = dossier
            # Le poste choisi à la validation est aussi enregistré sur la
            # réclamation (sinon « Poste souhaité » resterait vide).
            reclamation.poste = poste
            champs = ['statut', 'traite_par', 'traite_le', 'dossier_cree', 'poste']
            if derogation:
                reclamation.motif = f'Validé par dérogation (critères non remplis) : {derogation}'
                champs.append('motif')
            reclamation.save(update_fields=champs)
            # Enregistre la grille telle que cochée (audit, avec copie du libellé).
            self._enregistrer_grille(reclamation, actifs, coches, request.user)

        # Aucun email pendant le traitement : la personne (désormais RETENUE via
        # le dossier créé) sera notifiée à la publication de la liste.
        return Response(ReclamationAdminSerializer(reclamation).data)

    @action(detail=True, methods=['post'])
    def rejeter(self, request, pk=None):
        """Rejette la réclamation (motif obligatoire)."""
        reclamation = self.get_object()
        self._verifier_peut_decider(request.user, reclamation)
        motif = (request.data.get('motif') or '').strip()
        if not motif:
            raise ValidationError({'motif': "Un motif est obligatoire."})

        with transaction.atomic():
            # Verrou + re-contrôle sous verrou (cf. `valider`) : un rejet ne
            # peut pas écraser une validation simultanée, et inversement.
            reclamation = (
                ReclamationEligibilite.objects
                .select_for_update().get(pk=reclamation.pk)
            )
            if reclamation.statut != ReclamationEligibilite.Statut.EN_ATTENTE:
                raise ValidationError("Cette réclamation a déjà été traitée.")
            reclamation.statut = ReclamationEligibilite.Statut.REJETEE
            reclamation.motif = motif
            reclamation.traite_par = request.user
            reclamation.traite_le = timezone.now()
            reclamation.save(update_fields=['statut', 'motif', 'traite_par', 'traite_le'])
            # Trace la grille (ce que la personne a / n'a pas) même au rejet.
            self._enregistrer_grille(
                reclamation, self._criteres_reclamation(),
                self._ids_coches(request), request.user,
            )

        # Aucun email pendant le traitement (un rejet de réclamation ne notifie pas).
        return Response(ReclamationAdminSerializer(reclamation).data)

    @action(detail=True, methods=['post'])
    def rouvrir(self, request, pk=None):
        """Rouvre une réclamation DÉCIDÉE (validée/rejetée) → EN ATTENTE, pour
        correction (admin / superviseur uniquement).

        Si elle avait été VALIDÉE, le dossier créé est annulé (passé à REJETÉ,
        avec trace dans l'historique) afin de retirer la personne des retenus.
        L'agent re-traite ensuite normalement (Valider / Rejeter) avec le bon
        statut et le bon motif."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        with transaction.atomic():
            reclamation = (
                ReclamationEligibilite.objects
                .select_for_update().get(pk=self.get_object().pk)
            )
            if reclamation.statut == ReclamationEligibilite.Statut.EN_ATTENTE:
                raise ValidationError("Cette réclamation est déjà en attente.")
            # Annule le dossier créé à la validation (le retire des retenus).
            dossier = reclamation.dossier_cree
            if dossier and dossier.statut != Dossier.Statut.REJETE:
                ancien = dossier.statut
                dossier.statut = Dossier.Statut.REJETE
                dossier.save(update_fields=['statut', 'modifie_le'])
                HistoriqueStatut.objects.create(
                    dossier=dossier, ancien_statut=ancien,
                    nouveau_statut=Dossier.Statut.REJETE, par=request.user,
                    motif='Réclamation rouverte (correction) — dossier annulé',
                )
            reclamation.statut = ReclamationEligibilite.Statut.EN_ATTENTE
            reclamation.motif = ''
            reclamation.traite_par = None
            reclamation.traite_le = None
            reclamation.dossier_cree = None
            reclamation.save(update_fields=[
                'statut', 'motif', 'traite_par', 'traite_le', 'dossier_cree',
            ])
        return Response(ReclamationAdminSerializer(reclamation).data)

    # --- Répartition de la charge entre agents -------------------------

    # Réclamations « en attente » = en cours de traitement ; les autres
    # (validée / rejetée) sont déjà décidées (révision = supervision).
    STATUTS_EN_COURS = {ReclamationEligibilite.Statut.EN_ATTENTE}

    @action(detail=False, methods=['post'])
    def repartir(self, request):
        """Répartit équitablement des réclamations entre des agents (supervision).

        Corps : {agents: [id, …], statut?, appel?, q?, seulement_non_affectees?}.
        - `statut` : la catégorie filtrée à répartir (défaut : en attente). On
          peut répartir aussi les VALIDÉES / REJETÉES pour révision.
        - `seulement_non_affectees` (défaut True) : à False = RÉÉQUILIBRAGE
          (réaffecte aussi les déjà affectées pour équilibrer la charge).

        Éligibilité des agents selon la catégorie : « en attente » → agents de
        traitement (validateur/superviseur) ; catégorie DÉJÀ DÉCIDÉE → seuls les
        SUPERVISEURS (révision). Round-robin (parts égales ±1). Additif."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        agent_ids = request.data.get('agents') or []
        if not isinstance(agent_ids, list) or not agent_ids:
            raise ValidationError({'agents': "Sélectionnez au moins un agent."})

        statuts = [s for s in str(request.data.get('statut') or '').split(',') if s]
        if not statuts:
            statuts = [ReclamationEligibilite.Statut.EN_ATTENTE]
        # Catégorie déjà décidée si un statut hors « en cours » est visé.
        cible_decidee = any(s not in self.STATUTS_EN_COURS for s in statuts)
        eligible = roles.peut_superviser if cible_decidee else roles.peut_traiter

        User = get_user_model()
        trouves = {u.id: u for u in User.objects.filter(id__in=agent_ids, is_active=True)}
        agents = [trouves[i] for i in agent_ids if i in trouves and eligible(trouves[i])]
        if not agents:
            raise ValidationError({'agents': (
                "Pour une catégorie déjà décidée (validées / rejetées), seuls des "
                "superviseurs peuvent être affectés." if cible_decidee else
                "Aucun agent valide (validateur ou superviseur actif).")})

        qs = ReclamationEligibilite.objects.filter(statut__in=statuts)
        appel = request.data.get('appel')
        if appel:
            qs = qs.filter(appel_id=appel)
        for token in tokens_recherche(request.data.get('q') or ''):
            qs = qs.filter(texte_recherche__contains=token)
        if request.data.get('seulement_non_affectees', True):
            qs = qs.filter(affecte_a__isnull=True)
        ids = list(qs.order_by('cree_le').values_list('id', flat=True))

        par_agent = {u.id: [] for u in agents}
        for i, rid in enumerate(ids):
            par_agent[agents[i % len(agents)].id].append(rid)

        with transaction.atomic():
            for u in agents:
                lot = par_agent[u.id]
                if lot:
                    (ReclamationEligibilite.objects
                     .filter(id__in=lot).update(affecte_a=u))

        return Response({
            'total_reparti': len(ids),
            'par_agent': [
                {'agent_id': u.id, 'agent': (u.get_full_name() or u.email),
                 'attribuees': len(par_agent[u.id])}
                for u in agents
            ],
        })

    @action(detail=False, methods=['get'])
    def repartition(self, request):
        """Charge par agent : total affecté, en attente, traitées (back-office)."""
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        lignes = (
            ReclamationEligibilite.objects
            .filter(affecte_a__isnull=False)
            .values('affecte_a_id', 'affecte_a__first_name',
                    'affecte_a__last_name', 'affecte_a__email')
            .annotate(
                total=Count('id'),
                en_attente=Count('id', filter=Q(statut=ReclamationEligibilite.Statut.EN_ATTENTE)),
            )
            .order_by('affecte_a__first_name', 'affecte_a__last_name')
        )
        resultat = []
        for l in lignes:
            nom = f"{l['affecte_a__first_name']} {l['affecte_a__last_name']}".strip()
            resultat.append({
                'agent_id': l['affecte_a_id'],
                'agent': nom or l['affecte_a__email'],
                'total': l['total'],
                'en_attente': l['en_attente'],
                'traitees': l['total'] - l['en_attente'],
            })
        non_affectees = (
            ReclamationEligibilite.objects
            .filter(affecte_a__isnull=True,
                    statut=ReclamationEligibilite.Statut.EN_ATTENTE)
            .count()
        )
        return Response({'par_agent': resultat, 'non_affectees': non_affectees})

    # --- Emails (best-effort : n'annulent jamais l'opération) ----------

    def _accuse_reception(self, reclamation):
        try:
            envoyer_email(
                destinataire=reclamation.email,
                sujet='Accusé de réception de votre réclamation',
                template='reclamation_accuse.html',
                contexte={
                    'nom': f'{reclamation.nom} {reclamation.postnom} {reclamation.prenom}'.strip(),
                    'appel': reclamation.appel.titre,
                },
            )
        except Exception:  # noqa: BLE001
            pass

    def _notifier(self, reclamation, validee):
        sujet = ('Votre réclamation a été acceptée' if validee
                 else 'Décision concernant votre réclamation')
        template = 'reclamation_validee.html' if validee else 'reclamation_rejetee.html'
        try:
            envoyer_email(
                destinataire=reclamation.email,
                sujet=sujet,
                template=template,
                contexte={
                    'nom': f'{reclamation.nom} {reclamation.postnom} {reclamation.prenom}'.strip(),
                    'appel': reclamation.appel.titre,
                    'motif': reclamation.motif,
                },
            )
        except Exception:  # noqa: BLE001
            pass


class RapportsView(APIView):
    """Tableau de bord statistique (back-office) : éligibles, dossiers reçus,
    réclamations, niveaux de traitement, retenus par poste/origine.

    `?appel=<id>` restreint dossiers/réclamations à un appel (les éligibles
    restent globaux). Lecture seule, réservé au back-office.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")

        appel_id = request.query_params.get('appel')
        dossiers = Dossier.objects.all()
        reclamations = ReclamationEligibilite.objects.all()
        appel_obj = None
        if appel_id:
            dossiers = dossiers.filter(appel_id=appel_id)
            reclamations = reclamations.filter(appel_id=appel_id)
            appel_obj = AppelCandidature.objects.filter(pk=appel_id).first()

        # --- Éligibles (référentiel global, indépendant de l'appel) ---
        elig = ListeEligibilite.objects.all()
        eligibles = {
            'total': elig.count(),
            'publies': elig.filter(est_publie=True).count(),
            'par_type': {
                (row['type_eligibilite'] or 'autre'): row['n']
                for row in elig.values('type_eligibilite').annotate(n=Count('id'))
            },
        }

        # --- Dossiers ---
        # Les statistiques ne comptent PAS les brouillons (non soumis) : on se
        # base uniquement sur les dossiers déposés (et au-delà).
        soumis = dossiers.exclude(statut=Dossier.Statut.BROUILLON)  # « reçus »
        ds_par_statut = {
            row['statut']: row['n']
            for row in soumis.values('statut').annotate(n=Count('id'))
        }
        # Ventilation par statut des SEULS vrais dépôts en ligne (deposant
        # renseigné) : permet de savoir, parmi ces dépôts, combien retenus /
        # rejetés — sans le bruit des ajouts par validation (réclamation).
        ds_par_statut_en_ligne = {
            row['statut']: row['n']
            for row in soumis.filter(deposant__isnull=False)
                             .values('statut').annotate(n=Count('id'))
        }
        # Provenance fiable via `deposant` (jamais modifié) : vrais dépôts en
        # ligne (deposant renseigné) vs ajouts par validation de réclamation
        # (deposant=None). Le lien dossier_cree est effacé à l'annulation, donc
        # inutilisable ici (cf. get_queryset / provenance).
        nb_soumis = soumis.count()
        cat_en_ligne = soumis.filter(deposant__isnull=False).count()
        val_null = soumis.filter(deposant__isnull=True).annotate(
            _lie=Exists(ReclamationEligibilite.objects.filter(dossier_cree=OuterRef('pk')))
        )
        cat_validation = val_null.filter(_lie=True).count()
        cat_annule = val_null.filter(_lie=False).count()
        retenus_qs = dossiers.filter(statut=Dossier.Statut.RETENU)
        nb_ret = retenus_qs.count()
        nb_ret_reclam = retenus_qs.filter(
            Exists(ReclamationEligibilite.objects.filter(dossier_cree=OuterRef('pk')))
        ).count()

        par_poste = [
            {
                'poste': row['poste__libelle'] or 'Non précisé',
                'recus': row['recus'], 'retenus': row['retenus'],
            }
            for row in soumis.values('poste__libelle').annotate(
                recus=Count('id'),
                retenus=Count('id', filter=Q(statut=Dossier.Statut.RETENU)),
            ).order_by('-recus')
        ]

        # --- Réclamations ---
        rec_par_statut = {
            row['statut']: row['n']
            for row in reclamations.values('statut').annotate(n=Count('id'))
        }

        # --- Niveaux de traitement ---
        def g(d, k):
            return d.get(k, 0)
        ds_traite = (g(ds_par_statut, 'retenu') + g(ds_par_statut, 'non_retenu')
                     + g(ds_par_statut, 'rejete'))
        ds_attente = g(ds_par_statut, 'depose') + g(ds_par_statut, 'en_examen')
        rec_traite = g(rec_par_statut, 'validee') + g(rec_par_statut, 'rejetee')
        rec_attente = g(rec_par_statut, 'en_attente')

        # --- Recours ---
        # Rattachés à l'appel via leur source (dossier OU réclamation).
        recours_qs = Recours.objects.all()
        if appel_id:
            recours_qs = recours_qs.filter(
                Q(dossier__appel_id=appel_id) | Q(reclamation__appel_id=appel_id)
            )
        rc_par_statut = {
            row['statut']: row['n']
            for row in recours_qs.values('statut').annotate(n=Count('id'))
        }
        # « Traité » legacy compté avec les validés (cf. stats recours).
        rc_valide = g(rc_par_statut, 'valide') + g(rc_par_statut, 'traite')
        rc_rejete = g(rc_par_statut, 'rejete')
        rc_attente = g(rc_par_statut, 'en_attente')
        rc_total = rc_valide + rc_rejete + rc_attente
        rc_par_source = {
            'dossier': recours_qs.filter(dossier__isnull=False).count(),
            'reclamation': recours_qs.filter(reclamation__isnull=False).count(),
        }

        return Response({
            'appel': ({'id': appel_obj.id, 'titre': appel_obj.titre,
                       'liste_retenus_publiee': appel_obj.liste_retenus_publiee}
                      if appel_obj else None),
            'eligibles': eligibles,
            'dossiers': {
                'total': sum(ds_par_statut.values()),
                'recus': nb_soumis,
                'en_ligne': cat_en_ligne,
                'par_categorie': {
                    'en_ligne': cat_en_ligne,
                    'validation': cat_validation,
                    'annule': cat_annule,
                },
                'par_statut': ds_par_statut,
                'par_statut_en_ligne': ds_par_statut_en_ligne,
                'par_poste': par_poste,
            },
            'reclamations': {
                'total': sum(rec_par_statut.values()),
                'par_statut': rec_par_statut,
            },
            'recours': {
                'total': rc_total,
                'en_attente': rc_attente,
                'valide': rc_valide,
                'rejete': rc_rejete,
                'par_source': rc_par_source,
            },
            'traitement': {
                'dossiers': {'traite': ds_traite, 'en_attente': ds_attente,
                             'total': ds_traite + ds_attente},
                'reclamations': {'traite': rec_traite, 'en_attente': rec_attente,
                                 'total': rec_traite + rec_attente},
                'recours': {'traite': rc_valide + rc_rejete, 'en_attente': rc_attente,
                            'total': rc_total},
            },
            'retenus': {
                'total': nb_ret,
                'par_origine': {'reclamation': nb_ret_reclam,
                                'en_ligne': nb_ret - nb_ret_reclam},
                'par_poste': [{'poste': p['poste'], 'n': p['retenus']}
                              for p in par_poste if p['retenus']],
            },
        })


class RecoursThrottle(AnonRateThrottle):
    """Limite le formulaire public de recours (anti-spam)."""

    scope = 'reclamation'


class RecoursViewSet(viewsets.ModelViewSet):
    """Recours d'une personne (après publication des résultats).

    - **rechercher** : publique — par nom/postnom/prénom, renvoie distinctement
      les correspondances dans les réclamations ET les dossiers soumis (pour
      confirmer que la personne existe) ;
    - **création** : publique (identité + date de naissance + email + message),
      throttlée ;
    - **liste / détail / traiter** : réservés au back-office.
    """

    pagination_class = PaginationStandard
    parser_classes = [JSONParser]

    # Champs autorisés au tri (allowlist).
    TRI_AUTORISE = {'id', 'nom', 'statut', 'cree_le'}

    def get_serializer_class(self):
        if self.action == 'create':
            return RecoursCreationSerializer
        if self.action in ('update', 'partial_update'):
            return RecoursModificationSerializer
        return RecoursAdminSerializer

    def get_permissions(self):
        if self.action in ('create', 'rechercher'):
            return [AllowAny()]
        return [IsAuthenticated()]

    def update(self, request, *args, **kwargs):
        # Édition d'un recours (identité, contact, message, date de réception) :
        # réservée aux ADMINISTRATEURS (correction de saisie sensible).
        if not roles.est_admin(request.user):
            raise PermissionDenied("La modification d'un recours est réservée aux administrateurs.")
        response = super().update(request, *args, **kwargs)
        # Renvoie la vue complète (source, statut…) après modification.
        return Response(RecoursAdminSerializer(self.get_object()).data)

    def get_throttles(self):
        return [RecoursThrottle()] if self.action == 'create' else []

    def get_queryset(self):
        if not roles.acces_backoffice(self.request.user):
            return Recours.objects.none()
        qs = Recours.objects.select_related(
            'traite_par', 'affecte_a', 'dossier', 'dossier__appel', 'dossier__poste',
            'reclamation', 'reclamation__appel', 'reclamation__poste',
        )
        statut = self.request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        # Filtre d'affectation : « moi » = mon lot, « aucune » = non affectés,
        # un id = le lot d'un agent (pour la supervision de la répartition).
        affecte = self.request.query_params.get('affecte')
        if affecte == 'moi':
            qs = qs.filter(affecte_a=self.request.user)
        elif affecte == 'aucune':
            qs = qs.filter(affecte_a__isnull=True)
        elif affecte and affecte.isdigit():
            qs = qs.filter(affecte_a_id=int(affecte))
        q = (self.request.query_params.get('q') or '').strip()
        if q:
            qs = qs.filter(
                Q(nom__icontains=q) | Q(postnom__icontains=q)
                | Q(prenom__icontains=q) | Q(email__icontains=q)
            )
        ordering = self.request.query_params.get('ordering', '')
        if ordering.lstrip('-') in self.TRI_AUTORISE:
            qs = qs.order_by(ordering)
        else:
            qs = qs.order_by('-cree_le')
        return qs

    def create(self, request, *args, **kwargs):
        """Dépôt public d'un recours (réponse neutre)."""
        if not AppelCandidature.recours_ouverts():
            raise ValidationError(
                {'detail': "Le dépôt des recours est clôturé."}
            )
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(
            {'detail': "Votre recours a bien été enregistré. Il sera examiné par nos services."},
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['get'])
    def rechercher(self, request):
        """Public : cherche une identité (nom/postnom/prénom) et renvoie
        DISTINCTEMENT les correspondances dans les réclamations et dans les
        dossiers soumis (hors brouillon), pour que le demandeur reconnaisse et
        lie son enregistrement. Doublons (même identité) regroupés par source."""
        tokens = tokens_recherche(request.query_params.get('q', ''))
        if not tokens:
            return Response({'dossiers': [], 'reclamations': []})
        dossiers = (Dossier.objects
                    .exclude(statut=Dossier.Statut.BROUILLON)
                    .select_related('poste', 'appel'))
        reclams = ReclamationEligibilite.objects.select_related('poste', 'appel')
        for token in tokens:
            dossiers = dossiers.filter(texte_recherche__contains=token)
            reclams = reclams.filter(texte_recherche__contains=token)

        def dedupe(qs, source_type):
            """Une entrée par identité normalisée (le 1er enregistrement) :
            évite d'afficher plusieurs fois la même personne dans une source."""
            vus, out = set(), []
            for o in qs.order_by('nom', 'postnom', 'prenom', 'id'):
                cle = o.texte_recherche or f'{o.nom}|{o.postnom}|{o.prenom}'.lower()
                if cle in vus:
                    continue
                vus.add(cle)
                out.append({
                    'type': source_type, 'id': o.id,
                    'nom': o.nom, 'postnom': o.postnom, 'prenom': o.prenom,
                    'poste': o.poste.libelle if o.poste else None,
                    'appel': o.appel.titre,
                })
                if len(out) >= 25:
                    break
            return out

        return Response({
            'dossiers': dedupe(dossiers, 'dossier'),
            'reclamations': dedupe(reclams, 'reclamation'),
        })

    def _decider(self, request, statut_cible):
        """Tranche un recours (valider / rejeter) + note interne, et trace qui
        a décidé et quand. La décision reste INTERNE : elle n'actualise pas la
        liste publique des retenus ni le statut du dossier (la publication
        définitive est une étape ultérieure distincte)."""
        recours = self.get_object()
        if not roles.peut_traiter(request.user):
            raise PermissionDenied("Réservé aux administrateurs, superviseurs et validateurs.")
        if not roles.peut_decider_affecte(request.user, recours.affecte_a_id):
            raise PermissionDenied("Ce recours est affecté à un autre agent.")
        recours.statut = statut_cible
        reponse = request.data.get('reponse')
        if reponse is not None:
            recours.reponse = reponse.strip()
        recours.traite_par = request.user
        recours.traite_le = timezone.now()
        recours.save(update_fields=['statut', 'reponse', 'traite_par', 'traite_le'])
        return Response(RecoursAdminSerializer(recours).data)

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """Back-office : VALIDE le recours → rejoint la liste interne des
        « validés après recours » (distincte des autres validations)."""
        return self._decider(request, Recours.Statut.VALIDE)

    @action(detail=True, methods=['post'])
    def rejeter(self, request, pk=None):
        """Back-office : REJETTE le recours (décision défavorable)."""
        return self._decider(request, Recours.Statut.REJETE)

    @action(detail=True, methods=['post'])
    def domaine(self, request, pk=None):
        """Met à jour UNIQUEMENT le domaine (poste) du recours — réservé aux
        administrateurs. `poste_id` vide/absent = retour au domaine de la source.
        Si une entrée de liste définitive existe déjà pour ce recours (entrée
        figée), son domaine est aussi mis à jour pour rester cohérent."""
        if not roles.est_admin(request.user):
            raise PermissionDenied("La correction du domaine est réservée aux administrateurs.")
        recours = self.get_object()
        pid = request.data.get('poste_id')
        recours.poste = get_object_or_404(Poste, pk=pid) if pid else None
        recours.save(update_fields=['poste'])
        # Répercute sur la liste définitive déjà générée (entrées figées).
        libelle = _domaine_recours(recours)
        RetenuDefinitif.objects.filter(recours=recours).update(poste_libelle=libelle)
        return Response(RecoursAdminSerializer(recours).data)

    @action(detail=True, methods=['get'])
    def personne(self, request, pk=None):
        """Back-office : tous les enregistrements de la MÊME personne (doublons)
        — dossiers soumis + réclamations — avec leurs documents, pour examiner
        avant de décider. L'enregistrement source du recours est marqué
        (`est_source`). Correspondance tolérante par nom/postnom/prénom."""
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        recours = self.get_object()
        tokens = tokens_recherche(f'{recours.nom} {recours.postnom} {recours.prenom}')
        if not tokens:
            return Response({'dossiers': [], 'reclamations': []})
        dossiers = (Dossier.objects
                    .exclude(statut=Dossier.Statut.BROUILLON)
                    .select_related('poste', 'appel')
                    .prefetch_related('pieces__type_piece', 'historique'))
        reclams = (ReclamationEligibilite.objects
                   .select_related('poste', 'appel')
                   .prefetch_related('documents'))
        for token in tokens:
            dossiers = dossiers.filter(texte_recherche__contains=token)
            reclams = reclams.filter(texte_recherche__contains=token)

        rejetes_dossier = {Dossier.Statut.REJETE, Dossier.Statut.NON_RETENU}

        def _motif_rejet(d):
            """Motif de la décision défavorable (rejet/non-retenu) : porté par
            l'entrée d'historique qui a fait basculer le dossier dans cet état."""
            if d.statut not in rejetes_dossier:
                return ''
            for h in d.historique.all():   # prefetch, déjà trié -horodatage
                if h.nouveau_statut == d.statut and h.motif:
                    return h.motif
            return ''

        def _dossier(d):
            return {
                'type': 'dossier', 'id': d.id,
                'nom': d.nom, 'postnom': d.postnom, 'prenom': d.prenom,
                'email': d.email,
                'poste': d.poste.libelle if d.poste else None,
                'appel': d.appel.titre,
                'statut': d.statut, 'statut_libelle': d.get_statut_display(),
                'motif_rejet': _motif_rejet(d),
                'cree_le': d.cree_le,
                'est_source': d.id == recours.dossier_id,
                'documents': [{
                    'id': p.id,
                    'libelle': p.type_piece.libelle,
                    'nom_original': p.nom_original,
                    'url': f'/api/dossiers/{d.id}/pieces/{p.id}/telecharger/',
                } for p in d.pieces.all()],
            }

        def _reclam(r):
            return {
                'type': 'reclamation', 'id': r.id,
                'nom': r.nom, 'postnom': r.postnom, 'prenom': r.prenom,
                'email': r.email, 'telephone': r.telephone, 'message': r.message,
                'poste': r.poste.libelle if r.poste else None,
                'appel': r.appel.titre,
                'statut': r.statut, 'statut_libelle': r.get_statut_display(),
                'motif_rejet': (r.motif if r.statut == ReclamationEligibilite.Statut.REJETEE else ''),
                'cree_le': r.cree_le,
                'est_source': r.id == recours.reclamation_id,
                'documents': [{
                    'id': doc.id,
                    'libelle': doc.get_type_display(),
                    'nom_original': doc.nom_original,
                    'url': f'/api/reclamations/{r.id}/documents/{doc.id}/',
                } for doc in r.documents.all()],
            }

        return Response({
            'dossiers': [_dossier(d) for d in dossiers.order_by('-cree_le')],
            'reclamations': [_reclam(r) for r in reclams.order_by('-cree_le')],
        })

    @action(detail=True, methods=['post'], url_path='rouvrir')
    def rouvrir(self, request, pk=None):
        """Back-office : remet un recours traité en attente."""
        if not roles.peut_traiter(request.user):
            raise PermissionDenied("Réservé aux administrateurs, superviseurs et validateurs.")
        recours = self.get_object()
        recours.statut = Recours.Statut.EN_ATTENTE
        recours.traite_par = None
        recours.traite_le = None
        recours.save(update_fields=['statut', 'traite_par', 'traite_le'])
        return Response(RecoursAdminSerializer(recours).data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Compteurs par statut (pour les cartes KPI du back-office)."""
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        par_statut = {row['statut']: row['n']
                      for row in Recours.objects.values('statut').annotate(n=Count('id'))}
        # « Traités » legacy comptés avec les validés (décision favorable
        # implicite à l'époque) pour ne pas les perdre dans les compteurs.
        return Response({
            'total': sum(par_statut.values()),
            'en_attente': par_statut.get(Recours.Statut.EN_ATTENTE, 0),
            'valide': (par_statut.get(Recours.Statut.VALIDE, 0)
                       + par_statut.get(Recours.Statut.TRAITE, 0)),
            'rejete': par_statut.get(Recours.Statut.REJETE, 0),
        })

    @action(detail=False, methods=['post'])
    def repartir(self, request):
        """Répartit équitablement des recours entre des agents (supervision).

        Corps : {agents: [id, …], statut?, q?, seulement_non_affectes?}.
        - `statut` : la catégorie filtrée à répartir (défaut : en attente). On
          peut aussi répartir des recours DÉCIDÉS (validés/rejetés) pour révision.
        - `seulement_non_affectes` (défaut True) : à False = RÉÉQUILIBRAGE
          (réaffecte aussi les déjà affectés pour équilibrer la charge).

        Éligibilité des agents : « en attente » → agents de traitement
        (validateur/superviseur) ; catégorie DÉJÀ DÉCIDÉE → seuls les
        SUPERVISEURS (révision). Round-robin (parts égales ±1). Additif."""
        if not roles.peut_superviser(request.user):
            raise PermissionDenied("Réservé aux administrateurs et superviseurs.")
        agent_ids = request.data.get('agents') or []
        if not isinstance(agent_ids, list) or not agent_ids:
            raise ValidationError({'agents': "Sélectionnez au moins un agent."})

        statuts = [s for s in str(request.data.get('statut') or '').split(',') if s]
        if not statuts:
            statuts = [Recours.Statut.EN_ATTENTE]
        cible_decidee = any(s != Recours.Statut.EN_ATTENTE for s in statuts)
        eligible = roles.peut_superviser if cible_decidee else roles.peut_traiter

        trouves = {u.id: u for u in User.objects.filter(id__in=agent_ids, is_active=True)}
        agents = [trouves[i] for i in agent_ids if i in trouves and eligible(trouves[i])]
        if not agents:
            raise ValidationError({'agents': (
                "Pour une catégorie déjà décidée (validés / rejetés), seuls des "
                "superviseurs peuvent être affectés." if cible_decidee else
                "Aucun agent valide (validateur ou superviseur actif).")})

        qs = Recours.objects.filter(statut__in=statuts)
        q = (request.data.get('q') or '').strip()
        if q:
            qs = qs.filter(
                Q(nom__icontains=q) | Q(postnom__icontains=q)
                | Q(prenom__icontains=q) | Q(email__icontains=q)
            )
        if request.data.get('seulement_non_affectes', True):
            qs = qs.filter(affecte_a__isnull=True)
        lignes = list(qs.order_by('cree_le').values('id', 'nom', 'postnom', 'prenom'))

        # Regroupement par identité normalisée : tous les recours d'une MÊME
        # personne (doublons) restent assignés au MÊME agent — on ne les éclate
        # jamais entre plusieurs agents (facilite le traitement). Le round-robin
        # se fait donc sur les PERSONNES, pas sur les lignes (équilibrage par
        # personne, parts égales ±1 personne).
        groupes = {}
        ordre = []
        for row in lignes:
            cle = normaliser_texte(f"{row['nom']} {row['postnom']} {row['prenom']}")
            if cle not in groupes:
                groupes[cle] = []
                ordre.append(cle)
            groupes[cle].append(row['id'])

        par_agent = {u.id: [] for u in agents}
        for i, cle in enumerate(ordre):
            par_agent[agents[i % len(agents)].id].extend(groupes[cle])

        with transaction.atomic():
            for u in agents:
                lot = par_agent[u.id]
                if lot:
                    Recours.objects.filter(id__in=lot).update(affecte_a=u)

        return Response({
            'total_reparti': len(lignes),
            'personnes': len(ordre),
            'par_agent': [
                {'agent_id': u.id, 'agent': (u.get_full_name() or u.email),
                 'attribues': len(par_agent[u.id])}
                for u in agents
            ],
        })

    @action(detail=False, methods=['get'])
    def repartition(self, request):
        """Charge par agent : total affecté, en attente, traités (back-office)."""
        if not roles.acces_backoffice(request.user):
            raise PermissionDenied("Réservé au back-office.")
        lignes = (
            Recours.objects
            .filter(affecte_a__isnull=False)
            .values('affecte_a_id', 'affecte_a__first_name',
                    'affecte_a__last_name', 'affecte_a__email')
            .annotate(
                total=Count('id'),
                en_attente=Count('id', filter=Q(statut=Recours.Statut.EN_ATTENTE)),
            )
            .order_by('affecte_a__first_name', 'affecte_a__last_name')
        )
        resultat = []
        for l in lignes:
            nom = f"{l['affecte_a__first_name']} {l['affecte_a__last_name']}".strip()
            resultat.append({
                'agent_id': l['affecte_a_id'],
                'agent': nom or l['affecte_a__email'],
                'total': l['total'],
                'en_attente': l['en_attente'],
                'traites': l['total'] - l['en_attente'],
            })
        non_affectes = (
            Recours.objects
            .filter(affecte_a__isnull=True, statut=Recours.Statut.EN_ATTENTE)
            .count()
        )
        return Response({'par_agent': resultat, 'non_affectes': non_affectes})
