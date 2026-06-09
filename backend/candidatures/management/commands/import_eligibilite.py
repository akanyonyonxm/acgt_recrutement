"""Importe la liste d'éligibilité depuis un fichier Excel (.xlsx).

    python manage.py import_eligibilite chemin/vers/liste.xlsx
    python manage.py import_eligibilite liste.xlsx --vider --publier

Colonnes attendues (1re ligne = en-têtes, insensible à la casse/aux accents) :
    nom | postnom | prenom | type | annee | reference
Seul « nom » est obligatoire. « type » accepte « stage » ou « candidature ».
"""

from openpyxl import load_workbook

from django.core.management.base import BaseCommand, CommandError

from candidatures.models import ListeEligibilite
from candidatures.utils import normaliser_texte

# En-têtes reconnus (forme normalisée) -> champ du modèle.
COLONNES = {
    'nom': 'nom',
    'postnom': 'postnom',
    'post nom': 'postnom',
    'prenom': 'prenom',
    'type': 'type_eligibilite',
    'annee': 'annee',
    'reference': 'reference',
}


class Command(BaseCommand):
    help = "Importe la liste d'éligibilité depuis un fichier Excel."

    def add_arguments(self, parser):
        parser.add_argument('fichier', help='Chemin du fichier .xlsx')
        parser.add_argument('--vider', action='store_true',
                            help='Vide la liste existante avant import.')
        parser.add_argument('--publier', action='store_true',
                            help='Marque les lignes importées comme publiées.')

    def handle(self, *args, **options):
        try:
            classeur = load_workbook(options['fichier'], read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Fichier introuvable : {options['fichier']}")
        feuille = classeur.active

        lignes = feuille.iter_rows(values_only=True)
        try:
            entetes = next(lignes)
        except StopIteration:
            raise CommandError("Le fichier est vide.")

        # Associe chaque colonne du fichier à un champ du modèle.
        index = {}
        for i, entete in enumerate(entetes):
            champ = COLONNES.get(normaliser_texte(str(entete or '')))
            if champ:
                index[champ] = i
        if 'nom' not in index:
            raise CommandError(
                "Colonne « nom » introuvable. En-têtes attendus : "
                "nom, postnom, prenom, type, annee, reference."
            )

        if options['vider']:
            n = ListeEligibilite.objects.count()
            ListeEligibilite.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'{n} ligne(s) existante(s) supprimée(s).'))

        objets, ignorees = [], 0
        for ligne in lignes:
            valeur = lambda champ: (str(ligne[index[champ]]).strip()
                                    if champ in index and ligne[index[champ]] is not None else '')
            nom = valeur('nom')
            if not nom:
                ignorees += 1
                continue
            postnom, prenom = valeur('postnom'), valeur('prenom')
            objets.append(ListeEligibilite(
                nom=nom, postnom=postnom, prenom=prenom,
                type_eligibilite=self._type(valeur('type_eligibilite')),
                annee=self._annee(valeur('annee')),
                reference=valeur('reference'),
                est_publie=options['publier'],
                # texte_recherche calculé ici car bulk_create contourne save().
                texte_recherche=normaliser_texte(f'{nom} {postnom} {prenom}'),
            ))
        # Important (Windows surtout) : libère le verrou sur le fichier.
        classeur.close()

        ListeEligibilite.objects.bulk_create(objets, batch_size=500)
        self.stdout.write(self.style.SUCCESS(
            f'{len(objets)} personne(s) importée(s)'
            + (f', {ignorees} ligne(s) sans nom ignorée(s)' if ignorees else '')
            + (' et publiée(s)' if options['publier'] else '') + '.'
        ))

    @staticmethod
    def _type(valeur):
        v = normaliser_texte(valeur)
        if v.startswith('candidat') or v.startswith('emploi') or v.startswith('demande'):
            return ListeEligibilite.Type.CANDIDATURE
        return ListeEligibilite.Type.STAGE

    @staticmethod
    def _annee(valeur):
        chiffres = ''.join(c for c in valeur if c.isdigit())
        return int(chiffres) if chiffres else None
