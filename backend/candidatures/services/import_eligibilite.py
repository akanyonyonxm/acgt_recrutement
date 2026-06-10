"""Import de la liste d'éligibilité depuis un classeur Excel (.xlsx).

Logique partagée entre la commande CLI `import_eligibilite` et l'endpoint
d'upload du back-office (EligibiliteViewSet.importer). Renvoie un récapitulatif.

Colonnes attendues (1re ligne = en-têtes, insensible à la casse/aux accents) :
    code | nom | postnom | prenom | type | annee | reference
Seul « nom » est obligatoire. « type » accepte « stage » ou « candidature ».
Le « code » est un numéro public affiché sur la liste en ligne.
"""

from openpyxl import load_workbook

from candidatures.models import ListeEligibilite
from candidatures.utils import normaliser_texte

# En-têtes reconnus (forme normalisée) -> champ du modèle.
COLONNES = {
    'code': 'code',
    'numero': 'code',
    'nom': 'nom',
    'postnom': 'postnom',
    'post nom': 'postnom',
    'prenom': 'prenom',
    'type': 'type_eligibilite',
    'annee': 'annee',
    'reference': 'reference',
}


class ImportEligibiliteErreur(Exception):
    """Erreur fonctionnelle d'import (fichier vide/illisible, colonne « nom » absente)."""


def _type(valeur):
    v = normaliser_texte(valeur)
    if v.startswith('candidat') or v.startswith('emploi') or v.startswith('demande'):
        return ListeEligibilite.Type.CANDIDATURE
    return ListeEligibilite.Type.STAGE


def _annee(valeur):
    chiffres = ''.join(c for c in valeur if c.isdigit())
    return int(chiffres) if chiffres else None


def importer_eligibles(fichier, remplacer=False, publier=False):
    """Importe les éligibles depuis `fichier` (chemin OU objet fichier .xlsx).

    :param remplacer: vide la liste existante avant l'import (sinon : ajout).
    :param publier: marque les lignes importées comme publiées.
    :returns: dict {importes, ignorees, supprimes, publier}
    :raises ImportEligibiliteErreur: fichier vide/illisible ou colonne « nom » absente.
    """
    try:
        classeur = load_workbook(fichier, read_only=True, data_only=True)
    except FileNotFoundError as exc:
        raise ImportEligibiliteErreur("Fichier introuvable.") from exc
    except Exception as exc:  # openpyxl : fichier corrompu / mauvais format
        raise ImportEligibiliteErreur("Fichier illisible (format attendu : .xlsx).") from exc

    feuille = classeur.active
    lignes = feuille.iter_rows(values_only=True)
    try:
        entetes = next(lignes)
    except StopIteration:
        classeur.close()
        raise ImportEligibiliteErreur("Le fichier est vide.")

    # Associe chaque colonne du fichier à un champ du modèle.
    index = {}
    for i, entete in enumerate(entetes):
        champ = COLONNES.get(normaliser_texte(str(entete or '')))
        if champ:
            index[champ] = i
    if 'nom' not in index:
        classeur.close()
        raise ImportEligibiliteErreur(
            "Colonne « nom » introuvable. En-têtes attendus : "
            "code, nom, postnom, prenom, type, annee, reference."
        )

    objets, ignorees = [], 0
    for ligne in lignes:
        def valeur(champ):
            if champ in index and ligne[index[champ]] is not None:
                return str(ligne[index[champ]]).strip()
            return ''
        nom = valeur('nom')
        if not nom:
            ignorees += 1
            continue
        postnom, prenom = valeur('postnom'), valeur('prenom')
        objets.append(ListeEligibilite(
            nom=nom, postnom=postnom, prenom=prenom,
            code=valeur('code'),
            type_eligibilite=_type(valeur('type_eligibilite')),
            annee=_annee(valeur('annee')),
            reference=valeur('reference'),
            est_publie=publier,
            # texte_recherche calculé ici car bulk_create contourne save().
            texte_recherche=normaliser_texte(f'{nom} {postnom} {prenom}'),
        ))
    # Important (Windows surtout) : libère le verrou sur le fichier.
    classeur.close()

    supprimes = 0
    if remplacer:
        supprimes = ListeEligibilite.objects.count()
        ListeEligibilite.objects.all().delete()

    ListeEligibilite.objects.bulk_create(objets, batch_size=500)
    return {
        'importes': len(objets),
        'ignorees': ignorees,
        'supprimes': supprimes,
        'publier': publier,
    }
